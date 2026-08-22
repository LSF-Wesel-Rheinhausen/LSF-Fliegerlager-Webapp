from datetime import date, timedelta
from io import BytesIO
from xml.etree import ElementTree
from zipfile import ZipFile

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from billing.attendance_export import attendance_workbook_response
from billing.models import AttendanceDay, Participant, ParticipantFamilyMember
from billing.permissions import EDITOR_GROUP
from tests.factories import (
    CampFactory,
    GroupFactory,
    ParticipantFactory,
    ParticipantFamilyMemberFactory,
    SuperUserFactory,
    UserFactory,
)

pytestmark = pytest.mark.django_db


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
UNSAFE_FORMULA_PREFIXES = ("=", "+", "-", "@")


def create_attendance(*, person, day, status, comment=""):
    fields = {"date": day, "is_present": status == "present", "comment": comment}
    if isinstance(person, ParticipantFamilyMember):
        fields["participant"] = person.guardian
        fields["family_member"] = person
    else:
        fields["participant"] = person
    return AttendanceDay.objects.create(**fields)


@pytest.fixture
def attendance_dataset():
    camp = CampFactory(
        year=2026,
        starts_on=date(2026, 7, 1),
        ends_on=date(2026, 7, 5),
    )
    other_camp = CampFactory(
        year=2027,
        starts_on=date(2027, 7, 1),
        ends_on=date(2027, 7, 3),
    )
    participant = ParticipantFactory(
        camp=camp,
        first_name="=SUM(1,1)",
        last_name="Lovelace",
        email="private@example.test",
        phone="+49123456789",
        notes="private participant note",
        arrival_date=date(2026, 7, 2),
        departure_date=date(2026, 7, 4),
    )
    participant.birth_date = date(2010, 6, 30)
    participant.save()
    family_member = ParticipantFamilyMemberFactory(
        guardian=participant,
        first_name="Ada",
        last_name="=Family",
        role=ParticipantFamilyMember.Role.CHILD,
        arrival_date=date(2026, 7, 1),
        departure_date=date(2026, 7, 5),
    )
    family_member.birth_date = date(2015, 7, 2)
    family_member.save()
    legacy_participant = ParticipantFactory(
        camp=camp,
        first_name="Legacy",
        last_name="Dates",
        arrival_date=date(2026, 7, 1),
        departure_date=date(2026, 7, 3),
        attendance_tracking_enabled=False,
    )
    legacy_participant.birth_date = date(2012, 6, 30)
    legacy_participant.save()
    create_attendance(
        person=participant,
        day=date(2026, 7, 2),
        status="present",
        comment='+HYPERLINK("https://example.test")',
    )
    create_attendance(person=participant, day=date(2026, 7, 3), status="absent", comment="Krank")
    participant.attendance_tracking_enabled = True
    participant.save(update_fields=["attendance_tracking_enabled", "updated_at"])
    create_attendance(person=family_member, day=date(2026, 7, 2), status="absent", comment="Familiennotiz")
    family_member.attendance_tracking_enabled = True
    family_member.save(update_fields=["attendance_tracking_enabled", "updated_at"])
    ParticipantFactory(camp=other_camp, first_name="Other", last_name="Camp")

    archived = ParticipantFactory(
        camp=camp,
        first_name="Archived",
        last_name="Person",
        archived_at="2026-06-01T00:00:00Z",
    )
    cancelled = ParticipantFactory(
        camp=camp,
        first_name="Cancelled",
        last_name="Person",
        status=Participant.Status.CANCELLED,
    )
    inactive_guardian = ParticipantFactory(camp=camp, first_name="Inactive", last_name="Guardian")
    inactive_family_member = ParticipantFamilyMemberFactory(
        guardian=inactive_guardian,
        first_name="Inactive",
        last_name="Family",
        is_active=False,
    )
    return camp, participant, family_member, legacy_participant, archived, cancelled, inactive_family_member


def test_attendance_workbook_response_core_returns_workbook(attendance_dataset):
    camp, *_ = attendance_dataset

    response = attendance_workbook_response(camp)

    assert response["Content-Type"] == XLSX_CONTENT_TYPE
    assert response["Content-Disposition"] == 'attachment; filename="anwesenheit-2026.xlsx"'
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == ["Anwesenheit", "Tagesübersicht", "Legende"]


def test_attendance_workbook_requires_admin_and_allows_admin(client, attendance_dataset):
    camp, *_ = attendance_dataset
    url = reverse("attendance-workbook", args=[camp.pk])

    assert client.get(url).status_code == 302

    editor = UserFactory(username="editor")
    editor.groups.add(GroupFactory(name=EDITOR_GROUP))
    client.force_login(editor)
    assert client.get(url).status_code == 302

    client.force_login(UserFactory(username="ordinary"))
    assert client.get(url).status_code == 302

    client.force_login(SuperUserFactory(username="admin"))
    assert client.get(url).status_code == 200


def test_attendance_workbook_response_has_xlsx_mime_and_year_filename(client, attendance_dataset):
    camp, *_ = attendance_dataset

    response = attendance_workbook_response(camp)

    assert response["Content-Type"] == XLSX_CONTENT_TYPE
    assert response["Content-Disposition"] == 'attachment; filename="anwesenheit-2026.xlsx"'


def test_attendance_sheet_contains_sorted_matrix_and_typed_dates(client, attendance_dataset):
    (
        camp,
        participant,
        family_member,
        legacy_participant,
        archived,
        cancelled,
        inactive_family_member,
    ) = attendance_dataset
    workbook = load_workbook(
        BytesIO(attendance_workbook_response(camp).content),
        data_only=False,
    )
    sheet = workbook["Anwesenheit"]

    expected_days = [date(2026, 6, 27) + timedelta(days=offset) for offset in range(12)]
    assert [cell.value for cell in sheet[1]] == ["Name", "Alter", "Typ", *expected_days]
    assert all(isinstance(sheet.cell(1, column).value, date) for column in range(4, 16))
    assert all(sheet.cell(1, column).number_format != "General" for column in range(4, 16))
    assert [sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)] == [
        family_member.full_name,
        legacy_participant.full_name,
        "Inactive Guardian",
        "'" + participant.full_name,
    ]
    assert [sheet.cell(row, 2).value for row in range(2, 6)] == [10, 14, None, 16]
    assert [sheet.cell(row, 3).value for row in range(2, 6)] == [
        "Kind",
        "Teilnehmer",
        "Teilnehmer",
        "Teilnehmer",
    ]
    assert [sheet.cell(2, column).value for column in range(4, 16)] == [
        "–",
        "–",
        "–",
        "–",
        "AB",
        "AB",
        "AB",
        "AB",
        "–",
        "–",
        "–",
        "–",
    ]
    assert [sheet.cell(3, column).value for column in range(4, 16)] == [
        "–",
        "–",
        "–",
        "–",
        "AN",
        "AN",
        "–",
        "–",
        "–",
        "–",
        "–",
        "–",
    ]
    assert [sheet.cell(4, column).value for column in range(4, 16)] == ["–"] * 12
    assert [sheet.cell(5, column).value for column in range(4, 16)] == [
        "–",
        "–",
        "–",
        "–",
        "–",
        "AN",
        "AB",
        "–",
        "–",
        "–",
        "–",
        "–",
    ]
    exported_text = "\n".join(str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None)
    assert archived.full_name not in exported_text
    assert cancelled.full_name not in exported_text
    assert inactive_family_member.full_name not in exported_text
    assert "private@example.test" not in exported_text
    assert "+49123456789" not in exported_text
    assert "private participant note" not in exported_text
    assert "Krank" not in exported_text


def test_attendance_sheet_identifies_participant_child_and_companion_with_equal_names():
    camp = CampFactory(year=2026, starts_on=date(2026, 7, 1), ends_on=date(2026, 7, 3))
    participant = ParticipantFactory(
        camp=camp,
        first_name="Alex",
        last_name="Gleich",
        birth_date=date(1990, 1, 1),
        arrival_date=date(2026, 7, 1),
        departure_date=date(2026, 7, 3),
    )
    ParticipantFamilyMemberFactory(
        guardian=participant,
        first_name="Alex",
        last_name="Gleich",
        role=ParticipantFamilyMember.Role.CHILD,
        birth_date=date(2015, 1, 1),
        arrival_date=date(2026, 7, 1),
        departure_date=date(2026, 7, 3),
    )
    ParticipantFamilyMemberFactory(
        guardian=participant,
        first_name="Alex",
        last_name="Gleich",
        role=ParticipantFamilyMember.Role.COMPANION,
        birth_date=date(1988, 1, 1),
        arrival_date=date(2026, 7, 1),
        departure_date=date(2026, 7, 3),
    )

    workbook = load_workbook(BytesIO(attendance_workbook_response(camp).content), data_only=False)
    sheet = workbook["Anwesenheit"]
    summary = workbook["Tagesübersicht"]
    rows_by_type = {sheet.cell(row, 3).value: row for row in range(2, sheet.max_row + 1)}
    summary_rows = {summary.cell(row, 1).value: row for row in range(2, summary.max_row + 1)}

    assert [sheet.cell(1, column).value for column in range(1, 5)] == [
        "Name",
        "Alter",
        "Typ",
        date(2026, 6, 27),
    ]
    assert sheet.max_row == 4
    assert set(rows_by_type) == {"Teilnehmer", "Kind", "Begleitperson"}
    assert [sheet.cell(rows_by_type[person_type], 1).value for person_type in rows_by_type] == [
        "Alex Gleich",
        "Alex Gleich",
        "Alex Gleich",
    ]
    assert sheet.cell(rows_by_type["Teilnehmer"], 2).value == 36
    assert sheet.cell(rows_by_type["Kind"], 2).value == 11
    assert sheet.cell(rows_by_type["Begleitperson"], 2).value == 38
    assert all(not person_type.startswith(UNSAFE_FORMULA_PREFIXES) for person_type in rows_by_type)
    assert sheet.cell(1, 4).data_type == "d"
    assert sheet.freeze_panes == "D2"
    assert sheet.auto_filter.ref == "A1:M4"
    assert summary.cell(summary_rows[date(2026, 7, 1)], 2).value == 3
    assert summary.cell(summary_rows[date(2026, 7, 1)], 3).value == 0
    assert all(cell.data_type != "f" for row in summary.iter_rows() for cell in row)


def test_attendance_sheet_classifies_direct_participants_with_child_precedence():
    camp = CampFactory(year=2026, starts_on=date(2026, 7, 1), ends_on=date(2026, 7, 3))
    ParticipantFactory(camp=camp, first_name="Normal", last_name="A", is_child=False, is_companion=False)
    ParticipantFactory(camp=camp, first_name="Companion", last_name="B", is_child=False, is_companion=True)
    ParticipantFactory(camp=camp, first_name="Double", last_name="C", is_child=True, is_companion=True)
    ParticipantFactory(camp=camp, first_name="Child", last_name="D", is_child=True, is_companion=False)

    workbook = load_workbook(BytesIO(attendance_workbook_response(camp).content), data_only=False)
    sheet = workbook["Anwesenheit"]

    assert [sheet.cell(row, 1).value for row in range(2, 6)] == [
        "Normal A",
        "Companion B",
        "Double C",
        "Child D",
    ]
    assert [sheet.cell(row, 3).value for row in range(2, 6)] == [
        "Teilnehmer",
        "Begleitperson",
        "Kind",
        "Kind",
    ]


def test_attendance_summary_has_present_absent_totals_and_visible_comments(client, attendance_dataset):
    camp, *_ = attendance_dataset

    workbook = load_workbook(BytesIO(attendance_workbook_response(camp).content))
    sheet = workbook["Tagesübersicht"]

    assert [cell.value for cell in sheet[1]] == ["Datum", "Anwesend", "Abwesend", "Kommentare"]
    assert [sheet.cell(row, 1).value for row in range(2, 14)] == [
        date(2026, 6, 27) + timedelta(days=offset) for offset in range(12)
    ]
    assert [sheet.cell(row, 2).value for row in range(2, 14)] == [0, 0, 0, 0, 1, 2, 0, 0, 0, 0, 0, 0]
    assert [sheet.cell(row, 3).value for row in range(2, 14)] == [0, 0, 0, 0, 1, 1, 2, 1, 0, 0, 0, 0]
    assert sheet.cell(7, 4).value == '\'+HYPERLINK("https://example.test"); Familiennotiz'
    assert sheet.cell(8, 4).value == "Krank"
    assert isinstance(sheet.cell(2, 1).value, date)


def test_attendance_workbook_escapes_formula_like_names_and_comments(client, attendance_dataset):
    camp, *_ = attendance_dataset

    workbook = load_workbook(BytesIO(attendance_workbook_response(camp).content))
    values = [cell.value for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row]

    assert "=SUM(1,1)" not in values
    assert "=Family" not in values
    assert all(not isinstance(value, str) or not value.startswith(UNSAFE_FORMULA_PREFIXES) for value in values)


def test_attendance_workbook_uses_compact_colored_statuses_and_matching_status_logic(attendance_dataset):
    camp, participant, family_member, *_ = attendance_dataset

    workbook = load_workbook(BytesIO(attendance_workbook_response(camp).content))
    sheet = workbook["Anwesenheit"]
    legend = workbook["Legende"]
    date_columns = {sheet.cell(1, column).value: column for column in range(4, sheet.max_column + 1)}
    people_rows = {sheet.cell(row, 1).value: row for row in range(2, sheet.max_row + 1)}
    participant_row = people_rows["'" + participant.full_name]
    family_row = people_rows[family_member.full_name]

    expected = {
        "AN": "FFC6EFCE",
        "AB": "FFFFC7CE",
        "–": "FFE7E6E6",
    }
    status_cells = {
        "AN": sheet.cell(participant_row, date_columns[date(2026, 7, 2)]),
        # No AttendanceDay exists for this in-range family-member day.
        "AB": sheet.cell(family_row, date_columns[date(2026, 7, 1)]),
        # The participant's stay ends before this day.
        "–": sheet.cell(participant_row, date_columns[date(2026, 7, 4)]),
    }
    for status, cell in status_cells.items():
        assert cell.value == status
        assert cell.fill.fill_type == "solid"
        assert cell.fill.fgColor.rgb == expected[status]
        assert cell.alignment.horizontal == "center"

    assert list(legend.values) == [
        ("Status", "Bedeutung"),
        ("AN", "Anwesend"),
        ("AB", "Abwesend"),
        ("–", "Außerhalb des Aufenthaltszeitraums"),
    ]
    for row, status in enumerate(("AN", "AB", "–"), start=2):
        assert legend.cell(row, 1).fill.fgColor.rgb == expected[status]
    assert sheet.freeze_panes == "D2"
    assert sheet.auto_filter.ref == "A1:O5"
    assert sheet.column_dimensions["C"].width == 16
    assert sheet.column_dimensions["D"].width == 6
    assert workbook["Tagesübersicht"].freeze_panes == "A2"
    assert workbook["Tagesübersicht"].auto_filter.ref == "A1:D13"


def test_attendance_workbook_isolated_from_other_camps(client, attendance_dataset):
    camp, *_ = attendance_dataset

    workbook = load_workbook(BytesIO(attendance_workbook_response(camp).content))
    exported_text = "\n".join(
        str(cell.value) for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row if cell.value
    )

    assert "Other Camp" not in exported_text


def test_attendance_workbook_uses_guardian_stay_for_family_member_without_own_dates():
    camp = CampFactory(year=2026, starts_on=date(2026, 7, 1), ends_on=date(2026, 7, 3))
    guardian = ParticipantFactory(camp=camp, arrival_date=date(2026, 7, 1), departure_date=date(2026, 7, 3))
    member = ParticipantFamilyMemberFactory(
        guardian=guardian,
        role=ParticipantFamilyMember.Role.CHILD,
        arrival_date=None,
        departure_date=None,
    )

    workbook = load_workbook(BytesIO(attendance_workbook_response(camp).content))
    sheet = workbook["Anwesenheit"]
    member_row = next(row for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value == member.full_name)

    date_columns = {sheet.cell(1, column).value: column for column in range(4, sheet.max_column + 1)}
    assert [
        sheet.cell(member_row, date_columns[date(2026, 7, 1)]).value,
        sheet.cell(member_row, date_columns[date(2026, 7, 2)]).value,
    ] == ["AN", "AN"]


def test_attendance_workbook_preserves_unicode_names_and_age_boundaries():
    camp = CampFactory(year=2026, starts_on=date(2026, 7, 1), ends_on=date(2026, 7, 3))
    birthday = ParticipantFactory(
        camp=camp,
        first_name="Zoë",
        last_name="Ångström",
        birth_date=date(2010, 7, 1),
    )
    before_birthday = ParticipantFactory(
        camp=camp,
        first_name="Émile",
        last_name="Župan",
        birth_date=date(2010, 7, 2),
    )

    workbook = load_workbook(BytesIO(attendance_workbook_response(camp).content))
    sheet = workbook["Anwesenheit"]
    exported = {sheet.cell(row, 1).value: sheet.cell(row, 2).value for row in range(2, sheet.max_row + 1)}

    assert exported[birthday.full_name] == 16
    assert exported[before_birthday.full_name] == 15


def test_attendance_workbook_prints_two_week_matrix_one_page_wide_after_round_trip():
    camp = CampFactory(year=2026, starts_on=date(2026, 7, 1), ends_on=date(2026, 7, 15))
    ParticipantFactory(camp=camp, first_name="Alex", last_name="Print")

    response = attendance_workbook_response(camp)
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    sheet = workbook["Anwesenheit"]

    assert sheet.max_row == 2
    assert sheet.max_column == 25
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.page_setup.fitToHeight == 0
    assert sheet.page_setup.scale is None
    assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert str(sheet.print_area) == "'Anwesenheit'!$A$1:$Y$2"
    assert sheet.print_title_rows == "$1:$1"
    assert sheet.print_title_cols == "$A:$C"

    for unchanged_sheet in (workbook["Tagesübersicht"], workbook["Legende"]):
        assert unchanged_sheet.page_setup.orientation is None
        assert unchanged_sheet.page_setup.fitToWidth is None
        assert unchanged_sheet.page_setup.fitToHeight is None
        assert unchanged_sheet.page_setup.scale is None
        assert unchanged_sheet.sheet_properties.pageSetUpPr.fitToPage is None
        assert not unchanged_sheet.print_area
        assert unchanged_sheet.print_title_rows is None
        assert unchanged_sheet.print_title_cols is None

    with ZipFile(BytesIO(response.content)) as archive:
        worksheet_xml = ElementTree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
        workbook_xml = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    page_setup = worksheet_xml.find("x:pageSetup", namespace)
    page_setup_properties = worksheet_xml.find("x:sheetPr/x:pageSetUpPr", namespace)
    assert page_setup is not None
    assert page_setup.attrib == {
        "orientation": "landscape",
        "fitToHeight": "0",
        "fitToWidth": "1",
    }
    assert page_setup_properties is not None
    assert page_setup_properties.attrib["fitToPage"] == "1"
    defined_names = {
        element.attrib["name"]: element.text
        for element in workbook_xml.findall("x:definedNames/x:definedName", namespace)
        if element.attrib.get("localSheetId") == "0"
        and element.attrib["name"] in {"_xlnm.Print_Titles", "_xlnm.Print_Area"}
    }
    assert defined_names == {
        "_xlnm.Print_Titles": "'Anwesenheit'!$1:$1,'Anwesenheit'!$A:$C",
        "_xlnm.Print_Area": "'Anwesenheit'!$A$1:$Y$2",
    }


def test_attendance_workbook_handles_empty_camp_with_typed_day_columns():
    camp = CampFactory(year=2026, starts_on=date(2026, 7, 1), ends_on=date(2026, 7, 3))

    workbook = load_workbook(BytesIO(attendance_workbook_response(camp).content))
    sheet = workbook["Anwesenheit"]

    assert sheet.max_row == 1
    assert [cell.value for cell in sheet[1]][:3] == ["Name", "Alter", "Typ"]
    assert all(isinstance(cell.value, date) for cell in sheet[1][3:])
    assert sheet.cell(1, 4).value == date(2026, 6, 27)
    assert sheet.max_column == 13
    assert sheet.freeze_panes == "D2"
    assert sheet.auto_filter.ref == "A1:M1"
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.page_setup.fitToHeight == 0
    assert sheet.page_setup.scale is None
    assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert str(sheet.print_area) == "'Anwesenheit'!$A$1:$M$1"
    assert sheet.print_title_rows == "$1:$1"
    assert sheet.print_title_cols == "$A:$C"
    assert workbook["Tagesübersicht"].max_row == sheet.max_column - 2
    assert list(workbook["Legende"].values)[1:] == [
        ("AN", "Anwesend"),
        ("AB", "Abwesend"),
        ("–", "Außerhalb des Aufenthaltszeitraums"),
    ]


@pytest.mark.parametrize("missing_dates", [(None, None), (date(2026, 7, 1), None), (None, date(2026, 7, 5))])
def test_attendance_workbook_handles_missing_camp_dates_deterministically(client, missing_dates):
    camp = CampFactory(year=2026, starts_on=missing_dates[0], ends_on=missing_dates[1])

    response = attendance_workbook_response(camp)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Anwesenheit", "Tagesübersicht", "Legende"]
    assert [cell.value for cell in workbook["Anwesenheit"][1]] == ["Name", "Alter", "Typ"]
    assert [cell.value for cell in workbook["Tagesübersicht"][1]] == [
        "Datum",
        "Anwesend",
        "Abwesend",
        "Kommentare",
    ]
    assert list(workbook["Legende"].values)[1:] == [
        ("AN", "Anwesend"),
        ("AB", "Abwesend"),
        ("–", "Außerhalb des Aufenthaltszeitraums"),
    ]
