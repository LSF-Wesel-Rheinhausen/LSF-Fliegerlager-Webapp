from concurrent.futures import ThreadPoolExecutor
from datetime import date
from decimal import Decimal
from io import BytesIO
from threading import Barrier, Lock

import pytest
from django.contrib.auth import authenticate, get_user_model
from django.contrib.auth.hashers import check_password
from django.contrib.auth.models import Group
from django.core.management import CommandError, call_command
from django.db import close_old_connections
from django.urls import reverse
from openpyxl import load_workbook

from billing.management.commands.seed_local_test_db import Command
from billing.models import (
    AttendanceDay,
    Camp,
    CampKioskAccess,
    Charge,
    EmailDelivery,
    Expense,
    ExpenseAllocation,
    FirstAdminBootstrapLock,
    MealBookingOverride,
    MealOrder,
    MealPlanEntry,
    MealSignup,
    Participant,
    ParticipantBookingLink,
    ParticipantFamilyMember,
    ParticipantFamilyMemberPin,
    ParticipantPin,
    PasskeyCredential,
    Payment,
    PriceRule,
    PushSubscription,
    Settlement,
    SettlementRun,
)

User = get_user_model()


def seed() -> None:
    call_command("seed_local_test_db", verbosity=0)


@pytest.mark.django_db
def test_seed_creates_representative_personas_and_relationships():
    seed()

    active_camp = Camp.objects.get(name="Local Testlager Active", year=2026)
    pre_camp = Camp.objects.get(name="Local Testlager Pre-Camp", year=2027)
    archived_camp = Camp.objects.get(name="Local Testlager Archived", year=2025)

    assert active_camp.is_active is True
    assert pre_camp.is_active is False
    assert archived_camp.is_active is False
    assert set(Participant.objects.filter(camp=active_camp).values_list("first_name", flat=True)) >= {
        "AdultComplete",
        "ChildPartial",
        "InfantNoDob",
        "MissingDob",
    }
    assert Participant.objects.filter(camp=archived_camp, first_name="ArchivedSettled").exists()

    adult = Participant.objects.get(camp=active_camp, first_name="AdultComplete")
    child = Participant.objects.get(camp=active_camp, first_name="ChildPartial")
    infant = Participant.objects.get(camp=active_camp, first_name="InfantNoDob")
    missing_dob = Participant.objects.get(camp=active_camp, first_name="MissingDob")

    assert adult.birth_date == date(1990, 1, 2)
    assert child.is_child is True
    assert infant.birth_date == date(2024, 6, 1)
    assert missing_dob.birth_date is None
    assert adult.email == "adultcomplete@example.test"
    assert adult.phone
    assert missing_dob.email == ""
    assert missing_dob.phone == ""

    family = ParticipantFamilyMember.objects.get(guardian=adult, first_name="FamilyChild")
    companion = ParticipantFamilyMember.objects.get(guardian=adult, first_name="FamilyCompanion")
    assert family.role == ParticipantFamilyMember.Role.CHILD
    assert companion.role == ParticipantFamilyMember.Role.COMPANION
    assert companion.is_active is True
    assert ParticipantFamilyMember.objects.get(guardian=adult, first_name="InactiveFamily").is_active is False

    assert ParticipantBookingLink.objects.filter(status="accepted").exists()
    assert AttendanceDay.objects.filter(participant=adult, family_member__isnull=True).count() >= 4
    assert AttendanceDay.objects.filter(participant=adult, family_member=family).exists()


@pytest.mark.django_db
def test_seed_credentials_are_deterministic_but_hashed_and_cover_pin_states():
    seed()

    admin = authenticate(username="local-admin", password="LocalAdmin-417-Only!")
    editor = authenticate(username="local-editor", password="LocalEditor-417-Only!")
    assert admin is not None and admin.is_superuser is True
    assert editor is not None and editor.is_staff is False
    assert authenticate(username="local-inactive", password="LocalInactive-417-Only!") is None

    access = CampKioskAccess.objects.get(camp__name="Local Testlager Active")
    assert access.pin_hash != "864208"
    assert access.check_pin("864208") is True

    adult_pin = ParticipantPin.objects.get(participant__first_name="AdultComplete")
    unset_pin = ParticipantPin.objects.get(participant__first_name="MissingDob")
    locked_pin = ParticipantPin.objects.get(participant__first_name="ChildPartial")
    assert adult_pin.pin_hash != "2468"
    assert adult_pin.check_pin("2468") is True
    assert unset_pin.must_set_pin is True
    assert unset_pin.pin_hash == ""
    assert locked_pin.is_locked is True

    companion_pin = ParticipantFamilyMemberPin.objects.get(family_member__first_name="FamilyCompanion")
    assert companion_pin.pin_hash != "9753"
    assert companion_pin.check_pin("9753") is True
    assert check_password("LocalAdmin-417-Only!", admin.password)


@pytest.mark.django_db
def test_seed_is_idempotent_and_preserves_unrelated_records():
    seed()
    unrelated_camp = Camp.objects.create(name="Unrelated Local Camp", year=2098, is_active=False)
    unrelated_participant = Participant.objects.create(
        camp=unrelated_camp,
        first_name="Unrelated",
        last_name="Record",
    )

    model_counts = {
        model: model.objects.count()
        for model in (
            Camp,
            Participant,
            ParticipantFamilyMember,
            AttendanceDay,
            PriceRule,
            Charge,
            Payment,
            Expense,
            ExpenseAllocation,
            MealSignup,
            MealOrder,
            MealBookingOverride,
            MealPlanEntry,
            SettlementRun,
            Settlement,
        )
    }
    seed()

    assert {model: model.objects.count() for model in model_counts} == model_counts
    assert Participant.objects.filter(pk=unrelated_participant.pk).exists()
    assert Camp.objects.filter(pk=unrelated_camp.pk, name="Unrelated Local Camp").exists()


@pytest.mark.django_db
def test_seed_contains_issue_417_attendance_profile_and_export_prerequisites():
    seed()

    camp = Camp.objects.get(name="Local Testlager Active")
    adult = Participant.objects.get(camp=camp, first_name="AdultComplete")
    infant = Participant.objects.get(camp=camp, first_name="InfantNoDob")
    infant_attendance_dates = set(
        AttendanceDay.objects.filter(participant=infant, family_member__isnull=True).values_list("date", flat=True)
    )
    assert camp.starts_on - date.resolution * 4 in infant_attendance_dates
    assert camp.ends_on + date.resolution * 3 in infant_attendance_dates
    assert adult.arrival_date < camp.ends_on
    assert adult.departure_date > adult.arrival_date
    assert PriceRule.objects.filter(camp=camp).exists()
    assert Charge.objects.filter(participant=adult).exists()
    assert MealSignup.objects.filter(participant=adult).exists()
    assert MealSignup.objects.filter(variant=MealSignup.Variant.VEGAN_CHILD).exists()
    assert Settlement.objects.filter(participant=adult).exists()


@pytest.mark.django_db
def test_seed_financial_matrix_has_supported_states_without_refund_invention():
    seed()

    camp = Camp.objects.get(name="Local Testlager Active")
    adult = Participant.objects.get(camp=camp, first_name="AdultComplete")
    assert set(Charge.objects.filter(participant=adult).values_list("kind", flat=True)) >= {
        Charge.Kind.CAMP_FLAT,
        Charge.Kind.FOOD,
        Charge.Kind.DRINK,
        Charge.Kind.DONATION,
    }
    assert Charge.objects.filter(participant=adult, deleted_at__isnull=False).exists()
    assert Charge.objects.filter(participant=adult, deleted_at__isnull=True).exists()
    assert Payment.objects.filter(participant=adult, amount=Decimal("0.00")).exists()
    assert Payment.objects.filter(participant=adult, amount=Decimal("25.00")).exists()
    assert Payment.objects.filter(participant=adult, amount=Decimal("200.00")).exists()
    assert Expense.objects.filter(camp=camp, cost_center=Expense.CostCenter.FOOD_BREAKFAST).exists()
    assert ExpenseAllocation.objects.filter(expense__camp=camp).exists()
    assert SettlementRun.objects.filter(camp=camp).count() >= 2


@pytest.mark.django_db
def test_seed_does_not_create_outbound_or_authentication_side_effects():
    seed()

    assert EmailDelivery.objects.count() == 0
    assert PushSubscription.objects.count() == 0
    assert PasskeyCredential.objects.count() == 0


@pytest.mark.django_db
def test_seed_refuses_foreign_active_camp_without_partial_changes():
    foreign_camp = Camp.objects.create(name="Foreign Active Camp", year=2099)

    with pytest.raises(CommandError):
        seed()

    assert Camp.objects.filter(pk=foreign_camp.pk, is_active=True).exists()
    assert not User.objects.filter(username="local-admin").exists()


@pytest.mark.django_db
def test_seed_refuses_foreign_deterministic_user_collision_and_rolls_back():
    foreign_group = Group.objects.create(name="foreign-collision-role")
    foreign_user = User.objects.create_user(
        username="local-admin",
        email="foreign-owner@example.test",
        password="Foreign-Password-446!",
        first_name="Foreign",
        last_name="Owner",
        is_active=False,
        is_staff=False,
        is_superuser=False,
    )
    foreign_user.groups.add(foreign_group)

    with pytest.raises(CommandError):
        seed()

    foreign_user.refresh_from_db()
    assert foreign_user.email == "foreign-owner@example.test"
    assert foreign_user.first_name == "Foreign"
    assert foreign_user.last_name == "Owner"
    assert foreign_user.is_active is False
    assert foreign_user.is_staff is False
    assert foreign_user.is_superuser is False
    assert check_password("Foreign-Password-446!", foreign_user.password)
    assert list(foreign_user.groups.values_list("name", flat=True)) == ["foreign-collision-role"]
    assert Camp.objects.filter(name__startswith="Local Testlager").count() == 0
    assert not User.objects.filter(username="local-editor").exists()


@pytest.mark.django_db(transaction=True)
def test_concurrent_seed_invocations_are_serialized_without_partial_state(monkeypatch):
    FirstAdminBootstrapLock.objects.get_or_create(pk=1)
    barrier = Barrier(2)
    invocation_lock = Lock()
    invocation_count = 0
    original_acquire_seed_lock = Command._acquire_seed_lock

    def synchronized_acquire_seed_lock():
        nonlocal invocation_count
        with invocation_lock:
            invocation_count += 1
            should_wait = invocation_count <= 2
        if should_wait:
            barrier.wait(timeout=5)
        return original_acquire_seed_lock()

    monkeypatch.setattr(Command, "_acquire_seed_lock", staticmethod(synchronized_acquire_seed_lock))

    def invoke_seed():
        close_old_connections()
        try:
            call_command("seed_local_test_db", verbosity=0)
            return None
        except Exception as error:  # noqa: BLE001 - propagate ordinary worker failures to the test
            return error
        finally:
            close_old_connections()

    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(lambda _index: invoke_seed(), range(2)))

    assert results == [None, None]
    assert User.objects.filter(username__startswith="local-").count() == 4
    assert Camp.objects.filter(name__startswith="Local Testlager").count() == 3
    assert Participant.objects.filter(camp__name="Local Testlager Active").count() == 4


@pytest.mark.django_db
def test_seeded_issue_417_admin_views_profile_and_attendance_contracts(client):
    seed()
    camp = Camp.objects.get(name="Local Testlager Active")
    admin = User.objects.get(username="local-admin")
    adult = Participant.objects.get(camp=camp, first_name="AdultComplete")

    client.force_login(admin)
    detail = client.get(reverse("participant-detail", kwargs={"participant_id": adult.pk}))
    overview = client.get(reverse("camp-attendance-overview", kwargs={"camp_id": camp.pk}))

    assert detail.status_code == 200
    assert detail.context["participant"].email == "adultcomplete@example.test"
    assert detail.context["participant"].phone == "+49 000 1001"
    assert detail.context["participant"].birth_date == date(1990, 1, 2)
    assert {member.first_name for member in detail.context["family_members"]} >= {
        "FamilyChild",
        "FamilyCompanion",
    }
    assert {person["name"] for person in detail.context["attendance_people"]} >= {
        "AdultComplete Synthetic",
        "FamilyChild Synthetic",
    }

    assert overview.status_code == 200
    overview_people = {person["name"]: person for person in overview.context["attendance_people"]}
    adult_calendar = {entry["date"]: entry for entry in overview_people["AdultComplete Synthetic"]["calendar"]}
    child_calendar = {entry["date"]: entry for entry in overview_people["FamilyChild Synthetic"]["calendar"]}
    assert adult_calendar[date(2026, 8, 18)]["status"] == "present"
    assert adult_calendar[date(2026, 8, 19)]["status"] == "absent"
    assert {(comment["person"], comment["comment"]) for comment in overview.context["daily_comments"]} >= {
        ("AdultComplete Synthetic", "Seed attendance note")
    }
    assert child_calendar[date(2026, 8, 21)]["status"] == "present"
    assert child_calendar[date(2026, 8, 25)]["status"] == "disabled"


@pytest.mark.django_db
def test_seeded_issue_417_attendance_export_is_person_rows_by_day_columns(client):
    seed()
    camp = Camp.objects.get(name="Local Testlager Active")
    client.force_login(User.objects.get(username="local-admin"))

    response = client.get(reverse("attendance-workbook", kwargs={"camp_id": camp.pk}))

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    sheet = workbook["Anwesenheit"]
    summary = workbook["Tagesübersicht"]
    date_columns = {sheet.cell(1, column).value: column for column in range(3, sheet.max_column + 1)}
    people = {sheet.cell(row, 1).value: row for row in range(2, sheet.max_row + 1)}
    summary_rows = {summary.cell(row, 1).value: row for row in range(2, summary.max_row + 1)}

    assert sheet.max_row == 7
    assert sheet.max_column == 20
    assert {
        "AdultComplete Synthetic",
        "ChildPartial Synthetic",
        "FamilyChild Synthetic",
        "FamilyCompanion Synthetic",
    } <= people.keys()
    assert date(2026, 8, 14) in date_columns
    assert date(2026, 8, 31) in date_columns
    assert sheet.cell(people["AdultComplete Synthetic"], date_columns[date(2026, 8, 18)]).value == "Anwesend"
    assert sheet.cell(people["AdultComplete Synthetic"], date_columns[date(2026, 8, 19)]).value == "Abwesend"
    assert sheet.cell(people["ChildPartial Synthetic"], date_columns[date(2026, 8, 21)]).value == "Anwesend"
    assert sheet.cell(people["FamilyChild Synthetic"], date_columns[date(2026, 8, 21)]).value == "Anwesend"
    assert sheet.cell(people["FamilyChild Synthetic"], date_columns[date(2026, 8, 25)]).value == "Außerhalb"
    assert summary.cell(summary_rows[date(2026, 8, 19)], 4).value == "Seed attendance note"
