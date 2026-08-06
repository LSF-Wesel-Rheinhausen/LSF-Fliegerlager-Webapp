import csv
import re
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4

from billing.exporters import participant_pdf_response, settlement_snapshot_pdf_bytes
from billing.models import Charge, Expense, MealSignup
from billing.permissions import EDITOR_GROUP
from billing.services import create_settlement_run
from tests.factories import (
    CampFactory,
    ChargeFactory,
    DrinkEntryFactory,
    ExpenseFactory,
    GroupFactory,
    ParticipantFactory,
    PaymentFactory,
    SuperUserFactory,
    UserFactory,
)


class RecordingPdfCanvas:
    """Capture PDF drawing positions while replacing ReportLab's output boundary."""

    def __init__(self, *_args, **_kwargs):
        self.page_number = 1
        self.body_y_positions = []
        self.line_positions = []
        self.round_rect_positions = []
        self.text_positions = []

    def drawString(self, _x, y, text):
        self.body_y_positions.append(y)
        self.text_positions.append((self.page_number, y, text))

    def drawRightString(self, _x, y, text):
        self.body_y_positions.append(y)
        self.text_positions.append((self.page_number, y, text))

    def drawCentredString(self, _x, y, text):
        self.text_positions.append((self.page_number, y, text))

    def drawImage(self, _path, _x, y, **_kwargs):
        self.body_y_positions.append(y)

    def rect(self, _x, y, _width, _height, **_kwargs):
        self.body_y_positions.append(y)

    def roundRect(self, _x, y, _width, _height, **_kwargs):
        self.body_y_positions.append(y)
        self.round_rect_positions.append((self.page_number, y))

    def line(self, x1, y1, x2, y2):
        self.body_y_positions.extend((y1, y2))
        self.line_positions.append((self.page_number, x1, y1, x2, y2))

    def getPageNumber(self):
        return self.page_number

    def showPage(self):
        self.page_number += 1

    def save(self):
        return None

    def __getattr__(self, _name):
        return lambda *_args, **_kwargs: None


@pytest.fixture
def recording_pdf_canvases(monkeypatch):
    canvases = []

    def create_canvas(*args, **kwargs):
        recording_canvas = RecordingPdfCanvas(*args, **kwargs)
        canvases.append(recording_canvas)
        return recording_canvas

    monkeypatch.setattr("billing.exporters.canvas.Canvas", create_canvas)
    return canvases


@pytest.fixture
def export_dataset():
    camp = CampFactory(year=2026)
    participant = ParticipantFactory(
        camp=camp,
        first_name="Ada",
        last_name="Lovelace",
        email="ada@example.test",
        phone="01234",
        actual_nights=5,
    )
    ChargeFactory(
        participant=participant,
        kind=Charge.Kind.DRINK,
        description="Cola",
        quantity=Decimal("2.00"),
        unit_price=Decimal("2.50"),
    )
    ChargeFactory(
        participant=participant,
        kind=Charge.Kind.FOOD,
        description="Abendessen",
        quantity=Decimal("1.00"),
        unit_price=Decimal("7.00"),
    )
    DrinkEntryFactory(
        participant=participant,
        quantity=3,
        unit_price=Decimal("1.50"),
    )
    PaymentFactory(participant=participant, amount=Decimal("4.00"))
    ExpenseFactory(participant=participant, amount=Decimal("3.00"), status=Expense.Status.APPROVED)
    return camp, participant


def csv_rows(response):
    return list(csv.reader(StringIO(response.content.decode("utf-8"))))


def pdf_page_count(content):
    return len(re.findall(rb"/Type\s*/Page(?!s)", content))


@pytest.mark.django_db
@pytest.mark.parametrize(
    ("route_name", "arg_getter"),
    [
        ("export-settlements-csv", lambda camp, _participant: [camp.pk]),
        ("export-drinks-csv", lambda camp, _participant: [camp.pk]),
        ("export-workbook", lambda camp, _participant: [camp.pk]),
        ("participant-import-template", lambda camp, _participant: [camp.pk]),
        ("export-participant-pdf", lambda _camp, participant: [participant.pk]),
    ],
)
def test_export_routes_require_editor_access(client, export_dataset, route_name, arg_getter):
    camp, participant = export_dataset
    url = reverse(route_name, args=arg_getter(camp, participant))

    anonymous_response = client.get(url)

    assert anonymous_response.status_code == 302
    assert reverse("login") in anonymous_response["Location"]


@pytest.mark.django_db
@pytest.mark.parametrize(
    ("route_name", "arg_getter"),
    [
        ("export-settlements-csv", lambda camp, _participant: [camp.pk]),
        ("export-drinks-csv", lambda camp, _participant: [camp.pk]),
        ("export-workbook", lambda camp, _participant: [camp.pk]),
        ("participant-import-template", lambda camp, _participant: [camp.pk]),
        ("export-participant-pdf", lambda _camp, participant: [participant.pk]),
    ],
)
@pytest.mark.parametrize("user_kind", ["editor", "admin"])
def test_export_routes_allow_editor_and_admin_access(client, export_dataset, route_name, arg_getter, user_kind):
    camp, participant = export_dataset
    if user_kind == "admin":
        user = SuperUserFactory(username="admin")
    else:
        user = UserFactory(username="editor")
        user.groups.add(GroupFactory(name=EDITOR_GROUP))
    client.force_login(user)

    response = client.get(reverse(route_name, args=arg_getter(camp, participant)))

    assert response.status_code == 200


@pytest.mark.django_db
def test_settlement_csv_exports_calculated_kiosk_charges_payments_and_expenses(client, export_dataset):
    camp, _participant = export_dataset
    client.force_login(SuperUserFactory())

    response = client.get(reverse("export-settlements-csv", args=[camp.pk]))

    assert response.status_code == 200
    assert response["Content-Type"] == "text/csv; charset=utf-8"
    assert response["Content-Disposition"] == 'attachment; filename="abrechnung-2026.csv"'
    assert csv_rows(response) == [
        ["Nachname", "Vorname", "Brutto", "Förderung", "Soll", "Gezahlt", "Vorgestreckt", "Offen"],
        ["Lovelace", "Ada", "16.50", "0.00", "16.50", "4.00", "3.00", "9.50"],
    ]


@pytest.mark.django_db
def test_csv_exports_escape_formula_like_text_values(client):
    camp = CampFactory(year=2026)
    participant = ParticipantFactory(
        camp=camp,
        first_name="+SUM(1,1)",
        last_name="=cmd",
        actual_nights=1,
    )
    ChargeFactory(
        participant=participant,
        kind=Charge.Kind.DRINK,
        description="@malicious",
        quantity=Decimal("1.00"),
        unit_price=Decimal("2.00"),
    )
    client.force_login(SuperUserFactory())

    settlement_response = client.get(reverse("export-settlements-csv", args=[camp.pk]))
    drinks_response = client.get(reverse("export-drinks-csv", args=[camp.pk]))

    assert csv_rows(settlement_response)[1][:2] == ["'=cmd", "'+SUM(1,1)"]
    assert csv_rows(drinks_response)[1][:3] == ["'=cmd", "'+SUM(1,1)", "'@malicious"]


@pytest.mark.django_db
def test_csv_exports_leave_regular_text_values_unchanged(client, export_dataset):
    camp, _participant = export_dataset
    client.force_login(SuperUserFactory())

    response = client.get(reverse("export-drinks-csv", args=[camp.pk]))

    assert csv_rows(response)[2][:3] == ["Lovelace", "Ada", "Cola"]


@pytest.mark.django_db
def test_participant_import_template_export_contains_headers_and_examples(client, export_dataset):
    camp, _participant = export_dataset
    client.force_login(SuperUserFactory())

    response = client.get(reverse("participant-import-template", args=[camp.pk]))

    assert response.status_code == 200
    assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response["Content-Disposition"] == 'attachment; filename="teilnehmer_import_vorlage.xlsx"'

    wb = load_workbook(BytesIO(response.content))
    assert "Teilnehmer" in wb.sheetnames

    sheet = wb["Teilnehmer"]
    # Check headers
    headers = [cell.value for cell in sheet[1]]
    assert headers[:6] == ["Vorname*", "Nachname*", "Anreise*", "Abreise*", "Hilfssatz*", "Berufssatz*"]
    assert "Email" in headers
    assert "Notizen" in headers

    # Check that there is at least one example row
    assert sheet.max_row >= 2
    assert sheet.cell(row=2, column=1).value == "Max"
    assert sheet.cell(row=2, column=2).value == "Mustermann"


@pytest.mark.django_db
def test_drinks_csv_exports_legacy_entries_and_kiosk_drink_charges(client, export_dataset):
    camp, _participant = export_dataset
    client.force_login(SuperUserFactory())

    response = client.get(reverse("export-drinks-csv", args=[camp.pk]))

    rows = csv_rows(response)
    assert response.status_code == 200
    assert response["Content-Type"] == "text/csv; charset=utf-8"
    assert response["Content-Disposition"] == 'attachment; filename="getraenke-2026.csv"'
    assert rows[0] == ["Nachname", "Vorname", "Getränk", "Menge", "Einzelpreis", "Summe", "Erfasst am"]
    assert rows[1][:6] == ["Lovelace", "Ada", "Wasser", "3", "1.50", "4.50"]
    assert rows[2][:6] == ["Lovelace", "Ada", "Cola", "2.00", "2.50", "5.00"]


@pytest.mark.django_db
def test_workbook_export_contains_settlement_and_participant_sheets(client, export_dataset):
    camp, _participant = export_dataset
    client.force_login(SuperUserFactory())

    response = client.get(reverse("export-workbook", args=[camp.pk]))

    assert response.status_code == 200
    assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response["Content-Disposition"] == 'attachment; filename="fliegerlager-2026.xlsx"'

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["Abrechnung", "Teilnehmer", "Kostenstellen"]
    settlement_sheet = workbook["Abrechnung"]
    assert [cell.value for cell in settlement_sheet[1]] == [
        "Nachname",
        "Vorname",
        "Brutto",
        "Förderung",
        "Soll",
        "Gezahlt",
        "Vorgestreckt",
        "Offen",
    ]
    assert settlement_sheet["A2"].value == "Lovelace"
    assert settlement_sheet["B2"].value == "Ada"
    assert Decimal(str(settlement_sheet["C2"].value)) == Decimal("16.5")
    assert Decimal(str(settlement_sheet["H2"].value)) == Decimal("9.5")

    participants_sheet = workbook["Teilnehmer"]
    assert participants_sheet["A2"].value == "Lovelace"
    assert participants_sheet["B2"].value == "Ada"
    assert participants_sheet["C2"].value == "ada@example.test"
    assert participants_sheet["D2"].value == "01234"
    assert participants_sheet["H2"].value == 5


@pytest.mark.django_db
def test_workbook_export_compares_cost_center_income_and_expenses(client):
    camp = CampFactory(year=2026)
    participant = ParticipantFactory(camp=camp, first_name="Ada", last_name="Lovelace")
    charge = ChargeFactory(
        participant=participant,
        kind=Charge.Kind.FOOD,
        description="Frühstück Frühstück",
        quantity=Decimal("2.00"),
        unit_price=Decimal("4.00"),
    )
    MealSignup.objects.create(
        participant=participant,
        meal_date=date(2026, 7, 1),
        meal=MealSignup.Meal.BREAKFAST,
        variant=MealSignup.Variant.NORMAL,
        charge=charge,
    )
    ExpenseFactory(
        participant=participant,
        camp=camp,
        description="Brötchen",
        amount=Decimal("3.50"),
        status=Expense.Status.APPROVED,
        allocation_method=Expense.AllocationMethod.COST_CENTER,
        cost_center=Expense.CostCenter.FOOD_BREAKFAST,
    )
    client.force_login(SuperUserFactory())

    response = client.get(reverse("export-workbook", args=[camp.pk]))

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    sheet = workbook["Kostenstellen"]
    assert [cell.value for cell in sheet[1]] == [
        "Kostenstelle",
        "Einnahmen",
        "Ausgaben",
        "Saldo",
        "Anzahl Einnahmen",
        "Anzahl Ausgaben",
    ]
    assert [cell.value for cell in sheet[2]] == [
        "Unterkunft/Verpflegung - Frühstück",
        8,
        3.5,
        4.5,
        1,
        1,
    ]
    assert sheet["A4"].value == "Detaillierte Einnahmen pro Kostenstelle"
    assert sheet["A6"].value == "Unterkunft/Verpflegung - Frühstück"
    assert sheet["D6"].value == "Frühstück"
    assert sheet["E6"].value == 8
    assert sheet["A8"].value == "Detaillierte Ausgaben pro Kostenstelle"
    assert sheet["A10"].value == "Unterkunft/Verpflegung - Frühstück"
    assert sheet["D10"].value == "Brötchen"
    assert sheet["E10"].value == 3.5


@pytest.mark.django_db
def test_participant_pdf_export_returns_pdf_preview(client, export_dataset):
    _camp, participant = export_dataset
    client.force_login(SuperUserFactory())

    response = client.get(reverse("export-participant-pdf", args=[participant.pk]))

    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    assert response["Content-Disposition"] == f'inline; filename="abrechnung-{participant.pk}.pdf"'
    assert response["X-Frame-Options"] == "SAMEORIGIN"
    assert "frame-ancestors 'self'" in response["Content-Security-Policy"]
    assert response.content.startswith(b"%PDF-")


@pytest.mark.django_db
@pytest.mark.parametrize("invoice_source", ["current", "snapshot"])
def test_invoice_pdf_prints_every_grouped_booking_reference_and_date(
    invoice_source,
    recording_pdf_canvases,
):
    participant = ParticipantFactory()
    booking_date = date(2026, 7, 28)
    bookings = [
        ChargeFactory(
            participant=participant,
            kind=Charge.Kind.DRINK,
            description="Wasser (Kiosk)",
            occurred_on=booking_date,
        )
        for _index in range(12)
    ]

    if invoice_source == "snapshot":
        run = create_settlement_run(participant.camp, SuperUserFactory())
        settlement_snapshot_pdf_bytes(run.settlements.get())
    else:
        participant_pdf_response(participant)

    rendered_text = [text for _page, _y, text in recording_pdf_canvases[0].text_positions]
    assert "Datum: 28.07.2026" in rendered_text
    first_booking_line = "Buchungen: " + ", ".join(booking.booking_reference for booking in bookings[:6])
    assert first_booking_line in rendered_text
    for booking in bookings:
        assert any(booking.booking_reference in text for text in rendered_text)


@pytest.mark.django_db
@pytest.mark.parametrize("invoice_source", ["current", "snapshot"])
def test_invoice_pdf_places_separators_clear_of_the_following_position(invoice_source, recording_pdf_canvases):
    participant = ParticipantFactory()
    ChargeFactory(participant=participant, description="Erste Buchung", occurred_on=date(2026, 7, 28))
    ChargeFactory(participant=participant, description="Zweite Buchung", occurred_on=date(2026, 7, 29))

    if invoice_source == "snapshot":
        run = create_settlement_run(participant.camp, SuperUserFactory())
        settlement_snapshot_pdf_bytes(run.settlements.get())
    else:
        participant_pdf_response(participant)

    recording_canvas = recording_pdf_canvases[0]
    second_position_y = next(y for _page, y, text in recording_canvas.text_positions if text == "Zweite Buchung")
    first_separator_y = next(
        y1 for _page, x1, y1, x2, y2 in recording_canvas.line_positions if x1 == 50 and x2 == A4[0] - 50 and y1 == y2
    )

    assert first_separator_y - second_position_y >= 8


@pytest.mark.django_db
@pytest.mark.parametrize("invoice_source", ["current", "snapshot"])
def test_invoice_pdf_numbers_every_page(invoice_source, recording_pdf_canvases):
    participant = ParticipantFactory()
    for index in range(28):
        ChargeFactory(participant=participant, description=f"Buchung {index + 1}")

    if invoice_source == "snapshot":
        run = create_settlement_run(participant.camp, SuperUserFactory())
        settlement_snapshot_pdf_bytes(run.settlements.get())
    else:
        participant_pdf_response(participant)

    recording_canvas = recording_pdf_canvases[0]
    rendered_page_numbers = [
        (page, text) for page, _y, text in recording_canvas.text_positions if text.startswith("Seite ")
    ]

    assert rendered_page_numbers == [(page, f"Seite {page}") for page in range(1, recording_canvas.page_number)]


@pytest.mark.django_db
@pytest.mark.parametrize("invoice_source", ["current", "snapshot"])
def test_long_invoice_pdf_uses_continuation_pages_for_closing_blocks(invoice_source):
    participant = ParticipantFactory()
    for index in range(28):
        ChargeFactory(participant=participant, description=f"Buchung {index + 1}")

    if invoice_source == "snapshot":
        run = create_settlement_run(participant.camp, SuperUserFactory())
        content = settlement_snapshot_pdf_bytes(run.settlements.get())
    else:
        content = participant_pdf_response(participant).content

    assert pdf_page_count(content) >= 2


@pytest.mark.django_db
@pytest.mark.parametrize("invoice_source", ["current", "snapshot"])
def test_long_invoice_moves_summary_above_footer(invoice_source, recording_pdf_canvases):
    participant = ParticipantFactory()
    for index in range(28):
        ChargeFactory(participant=participant, description=f"Buchung {index + 1}")

    if invoice_source == "snapshot":
        run = create_settlement_run(participant.camp, SuperUserFactory())
        settlement_snapshot_pdf_bytes(run.settlements.get())
    else:
        participant_pdf_response(participant)

    recording_canvas = recording_pdf_canvases[0]
    summary_pages = [page for page, _y, text in recording_canvas.text_positions if text == "Brutto:"]
    footer_pages = [
        page
        for page, _y, text in recording_canvas.text_positions
        if text.startswith("Erstellt mit der Fliegerlagerabrechnung")
    ]

    final_page = max(footer_pages)
    assert summary_pages == [final_page]
    assert footer_pages == list(range(1, final_page + 1))
    assert all(y == 30 or y >= 55 for y in recording_canvas.body_y_positions)


@pytest.mark.django_db
@pytest.mark.parametrize("invoice_source", ["current", "snapshot"])
@pytest.mark.parametrize(
    ("balance_kind", "instruction_heading"),
    [
        ("debit", "Zahlungsinformationen"),
        ("credit", "Guthaben & Auszahlung"),
    ],
)
def test_long_invoice_keeps_payment_box_above_footer(
    invoice_source,
    balance_kind,
    instruction_heading,
    recording_pdf_canvases,
):
    participant = ParticipantFactory(camp=CampFactory(iban="DE02120300000000202051"))
    for index in range(20):
        ChargeFactory(participant=participant, description=f"Buchung {index + 1}")
    if balance_kind == "credit":
        PaymentFactory(participant=participant, amount=Decimal("210.00"))

    if invoice_source == "snapshot":
        run = create_settlement_run(participant.camp, SuperUserFactory())
        settlement_snapshot_pdf_bytes(run.settlements.get())
    else:
        participant_pdf_response(participant)

    recording_canvas = recording_pdf_canvases[0]
    instruction_pages = [page for page, _y, text in recording_canvas.text_positions if text == instruction_heading]
    footer_pages = [
        page
        for page, _y, text in recording_canvas.text_positions
        if text.startswith("Erstellt mit der Fliegerlagerabrechnung")
    ]

    final_page = max(footer_pages)
    assert instruction_pages == [final_page]
    assert footer_pages == list(range(1, final_page + 1))
    assert len(recording_canvas.round_rect_positions) == 1
    payment_page, payment_bottom = recording_canvas.round_rect_positions[0]
    assert payment_page == final_page
    assert payment_bottom >= 55
