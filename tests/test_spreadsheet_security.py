from datetime import date
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from django.core.exceptions import ValidationError
from openpyxl import Workbook, load_workbook

from billing.exporters import CSV_FORMULA_PREFIXES, _write_cost_center_sheet_from_snapshot, camp_workbook_response
from billing.forms import MealBookingForm
from billing.importers import preview_participants
from billing.models import Expense, MealSignup
from tests.factories import CampFactory, ExpenseFactory, ParticipantFactory


@pytest.mark.django_db
def test_xlsx_export_neutralizes_formula_like_cost_center_descriptions():
    camp = CampFactory(year=2026)
    participant = ParticipantFactory(camp=camp, first_name="Ada", last_name="Lovelace")
    ExpenseFactory(
        participant=participant,
        camp=camp,
        description="=1+1",
        amount="10.00",
        status=Expense.Status.APPROVED,
        allocation_method=Expense.AllocationMethod.COST_CENTER,
        cost_center=Expense.CostCenter.FOOD_BREAKFAST,
    )

    workbook_response = camp_workbook_response(camp)
    workbook = load_workbook(BytesIO(workbook_response.content), data_only=False)
    sheet = workbook["Kostenstellen"]
    description_cell = next(cell for row in sheet.iter_rows() for cell in row if cell.value in {"=1+1", "'=1+1"})

    assert description_cell.value == "'=1+1"
    assert description_cell.data_type == "s"


def test_xlsx_cost_center_sheet_neutralizes_every_formula_prefix():
    snapshot = [
        {
            "label": "=Kostenstelle",
            "income": "10.00",
            "expense_total": "5.00",
            "balance": "5.00",
            "income_count": 1,
            "expense_count": 1,
            "income_details": [
                {
                    "meal_date": "2026-08-01",
                    "participant_name": "+Teilnehmer",
                    "family_member_name": "-Familienmitglied",
                    "description": "@Beschreibung",
                    "amount": "10.00",
                }
            ],
            "expense_details": [
                {
                    "paid_date": "2026-08-02",
                    "applicant_name": "+Antragsteller",
                    "description": "-Beschreibung",
                    "amount": "5.00",
                }
            ],
        }
    ]
    sheet = Workbook().active

    _write_cost_center_sheet_from_snapshot(sheet, snapshot)

    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.lstrip().startswith(CSV_FORMULA_PREFIXES), cell.value
                assert cell.data_type == "s"


@pytest.mark.django_db
@pytest.mark.parametrize(
    ("starts_on", "ends_on"),
    [
        (None, None),
        (None, date(2026, 8, 10)),
        (date(2026, 8, 20), None),
        (date(2026, 8, 20), date(2026, 8, 10)),
    ],
)
def test_meal_booking_form_rejects_dates_when_camp_bounds_are_not_valid(starts_on, ends_on):
    camp = CampFactory(starts_on=starts_on, ends_on=ends_on)
    participant = ParticipantFactory(camp=camp)

    form = MealBookingForm(
        data={
            "meal_dates": ["2099-12-31"],
            "meal": MealSignup.Meal.DINNER,
            "variant": MealSignup.Variant.NORMAL,
        },
        participant=participant,
    )

    assert form.is_valid() is False
    assert "2099-12-31" not in [value for value, _label in form.fields["meal_dates"].choices]


def _xlsx_with_large_shared_string(size: int = 1_048_770) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["first_name", "last_name", "arrival_date", "departure_date", "hilfssatz", "berufssatz"])
    sheet.append(["A", "B", "01.08.2026", "02.08.2026", 1, 1])

    original = BytesIO()
    workbook.save(original)
    original.seek(0)
    rewritten = BytesIO()

    with ZipFile(original) as source, ZipFile(rewritten, "w", compression=ZIP_DEFLATED) as target:
        for info in source.infolist():
            content = source.read(info.filename)
            if info.filename == "[Content_Types].xml":
                content = content.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/sharedStrings.xml" '
                    b'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml" />'
                    b"</Types>",
                )
            elif info.filename == "xl/_rels/workbook.xml.rels":
                content = content.replace(
                    b"</Relationships>",
                    b'<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/'
                    b'sharedStrings" '
                    b'Target="sharedStrings.xml" Id="rId4" /></Relationships>',
                )
            elif info.filename == "xl/worksheets/sheet1.xml":
                content = content.replace(
                    b'<c r="A2" t="inlineStr"><is><t>A</t></is></c>',
                    b'<c r="A2" t="s"><v>0</v></c>',
                )
            target.writestr(info.filename, content)

        shared_string = "A" * size
        target.writestr(
            "xl/sharedStrings.xml",
            (
                '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'count="1" uniqueCount="1"><si><t>'
                f"{shared_string}"
                "</t></si></sst>"
            ).encode(),
        )

    rewritten.seek(0)
    return rewritten


def test_xlsx_preview_accepts_a_normal_workbook_after_archive_preflight():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["first_name", "last_name", "arrival_date", "departure_date", "hilfssatz", "berufssatz"])
    sheet.append(["Ada", "Lovelace", "01.08.2026", "02.08.2026", 1, 1])
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)

    rows = preview_participants(payload, "teilnehmer.xlsx")

    assert len(rows) == 1
    assert rows[0].valid is True
    assert rows[0].data["first_name"] == "Ada"


def test_xlsx_preview_rejects_shared_string_expansion_before_openpyxl_materializes_it(monkeypatch):
    payload = _xlsx_with_large_shared_string()
    with ZipFile(payload) as archive:
        shared_strings = archive.getinfo("xl/sharedStrings.xml")
        assert shared_strings.file_size > shared_strings.compress_size * 100
    assert len(payload.getbuffer()) < 10_000
    payload.seek(0)
    monkeypatch.setattr(
        "billing.importers.load_workbook",
        lambda *args, **kwargs: pytest.fail("openpyxl darf die abgewiesene XLSX-Datei nicht materialisieren"),
    )
    with pytest.raises(ValidationError, match="Kompression|Expansion|sicher gelesen"):
        preview_participants(payload, "teilnehmer.xlsx")


def test_xlsx_preview_rejects_excessive_archive_entries_before_openpyxl(monkeypatch):
    payload = BytesIO()
    with ZipFile(payload, "w", compression=ZIP_DEFLATED) as archive:
        for index in range(1001):
            archive.writestr(f"unused/{index}.xml", b"")
    payload.seek(0)
    monkeypatch.setattr(
        "billing.importers.load_workbook",
        lambda *args, **kwargs: pytest.fail("openpyxl darf das Eintrags-Flooding nicht materialisieren"),
    )

    with pytest.raises(ValidationError, match="Einträge"):
        preview_participants(payload, "teilnehmer.xlsx")
