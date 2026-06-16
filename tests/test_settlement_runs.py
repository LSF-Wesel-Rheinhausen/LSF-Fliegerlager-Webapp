from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth.models import Group
from django.db import IntegrityError
from django.urls import reverse
from openpyxl import load_workbook

from billing.models import Charge, Expense, MealSignup, Settlement, SettlementRun
from billing.permissions import EDITOR_GROUP
from billing.services import create_settlement_run
from tests.factories import ChargeFactory, ExpenseFactory, ParticipantFactory, SuperUserFactory, UserFactory


@pytest.mark.django_db
def test_create_settlement_run_versions_immutable_participant_snapshots():
    user = SuperUserFactory()
    participant = ParticipantFactory(first_name="Ada", last_name="Lovelace")
    charge = ChargeFactory(participant=participant, unit_price=Decimal("10.00"))

    first = create_settlement_run(participant.camp, user)
    first_snapshot = first.settlements.get()
    charge.unit_price = Decimal("20.00")
    charge.save(update_fields=["unit_price"])
    second = create_settlement_run(participant.camp, user)

    first_snapshot.refresh_from_db()
    assert first.version == 1
    assert second.version == 2
    assert first_snapshot.participant_name == "Ada Lovelace"
    assert first_snapshot.total_due == Decimal("10.00")
    assert second.settlements.get().total_due == Decimal("20.00")
    assert first_snapshot.data["lines"][0]["unit_price"] == "10.00"


@pytest.mark.django_db
def test_new_settlement_run_excludes_archived_participants():
    user = SuperUserFactory()
    active = ParticipantFactory()
    archived = ParticipantFactory(camp=active.camp, archived_at="2026-06-09T12:00:00Z")

    run = create_settlement_run(active.camp, user)

    assert list(run.settlements.values_list("participant_id", flat=True)) == [active.pk]
    assert not run.settlements.filter(participant=archived).exists()


@pytest.mark.django_db
def test_admin_can_create_and_view_settlement_run(client):
    user = SuperUserFactory()
    participant = ParticipantFactory()
    client.force_login(user)

    response = client.post(reverse("settlement-run-create", args=[participant.camp_id]))

    run = SettlementRun.objects.get()
    assert response.status_code == 302
    assert response["Location"] == reverse("settlement-run-detail", args=[run.pk])
    detail = client.get(response["Location"])
    assert detail.status_code == 200
    assert b"V1" in detail.content


@pytest.mark.django_db
def test_editor_can_view_but_not_create_settlement_run(client):
    admin = SuperUserFactory()
    participant = ParticipantFactory()
    run = create_settlement_run(participant.camp, admin)
    editor = UserFactory()
    editor.groups.add(Group.objects.create(name=EDITOR_GROUP))
    client.force_login(editor)

    detail = client.get(reverse("settlement-run-detail", args=[run.pk]))
    create = client.post(reverse("settlement-run-create", args=[participant.camp_id]))

    assert detail.status_code == 200
    assert create.status_code == 302
    assert SettlementRun.objects.count() == 1


@pytest.mark.django_db
def test_historical_exports_use_snapshot_data(client):
    user = SuperUserFactory()
    participant = ParticipantFactory(first_name="Ada", last_name="Lovelace")
    ChargeFactory(participant=participant, unit_price=Decimal("12.00"))
    run = create_settlement_run(participant.camp, user)
    snapshot = run.settlements.get()
    participant.first_name = "Changed"
    participant.save(update_fields=["first_name"])
    client.force_login(user)

    csv_response = client.get(reverse("settlement-run-csv", args=[run.pk]))
    workbook_response = client.get(reverse("settlement-run-workbook", args=[run.pk]))
    pdf_response = client.get(reverse("settlement-snapshot-pdf", args=[snapshot.pk]))

    assert csv_response.status_code == 200
    assert "Ada Lovelace" in csv_response.content.decode("utf-8")
    workbook = load_workbook(BytesIO(workbook_response.content), data_only=True)
    assert workbook["Abrechnung"]["A2"].value == "Ada Lovelace"
    assert pdf_response.status_code == 200
    assert pdf_response.content.startswith(b"%PDF-")


@pytest.mark.django_db
def test_historical_workbook_uses_cost_center_snapshot_data(client):
    user = SuperUserFactory()
    participant = ParticipantFactory(first_name="Ada", last_name="Lovelace")
    breakfast_charge = ChargeFactory(
        participant=participant,
        kind=Charge.Kind.FOOD,
        description="Frühstück Frühstück",
        unit_price=Decimal("4.00"),
    )
    MealSignup.objects.create(
        participant=participant,
        meal_date=date(2026, 7, 1),
        meal=MealSignup.Meal.BREAKFAST,
        variant=MealSignup.Variant.NORMAL,
        charge=breakfast_charge,
    )
    ExpenseFactory(
        participant=participant,
        camp=participant.camp,
        description="Brötchen",
        amount=Decimal("3.50"),
        status=Expense.Status.APPROVED,
        allocation_method=Expense.AllocationMethod.COST_CENTER,
        cost_center=Expense.CostCenter.FOOD_BREAKFAST,
    )
    run = create_settlement_run(participant.camp, user)

    dinner_charge = ChargeFactory(
        participant=participant,
        kind=Charge.Kind.FOOD,
        description="Abendessen Abendessen",
        unit_price=Decimal("7.00"),
    )
    MealSignup.objects.create(
        participant=participant,
        meal_date=date(2026, 7, 2),
        meal=MealSignup.Meal.DINNER,
        variant=MealSignup.Variant.NORMAL,
        charge=dinner_charge,
    )
    ExpenseFactory(
        participant=participant,
        camp=participant.camp,
        description="Nudeln",
        amount=Decimal("5.00"),
        status=Expense.Status.APPROVED,
        allocation_method=Expense.AllocationMethod.COST_CENTER,
        cost_center=Expense.CostCenter.FOOD_DINNER,
    )
    client.force_login(user)

    workbook_response = client.get(reverse("settlement-run-workbook", args=[run.pk]))

    workbook = load_workbook(BytesIO(workbook_response.content), data_only=True)
    cost_center_sheet = workbook["Kostenstellen"]
    assert cost_center_sheet["A2"].value == "Unterkunft/Verpflegung - Frühstück"
    assert cost_center_sheet["B2"].value == 4
    assert cost_center_sheet["C2"].value == 3.5
    values = [cell[0] for cell in cost_center_sheet.iter_rows(min_row=1, max_col=1, values_only=True)]
    assert "Unterkunft/Verpflegung - Abendessen" not in values


@pytest.mark.django_db
def test_settlement_snapshot_is_unique_per_run_and_participant():
    user = SuperUserFactory()
    participant = ParticipantFactory()
    run = create_settlement_run(participant.camp, user)

    with pytest.raises(IntegrityError):
        Settlement.objects.create(
            run=run,
            participant=participant,
            total_due=0,
            total_paid=0,
            total_advanced=0,
            balance=0,
        )
