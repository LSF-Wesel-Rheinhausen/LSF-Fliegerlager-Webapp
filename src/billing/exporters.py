import csv
from collections.abc import Iterable, Sequence
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any

from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from .models import Charge, DrinkEntry, Participant, Settlement, SettlementRun
from .services import calculate_camp_settlements, calculate_participant_settlement, get_cost_center_evaluation, money

CSV_FORMULA_PREFIXES = ("=", "+", "-", "@")
PDF_CONTENT_BOTTOM = 55
PDF_LINE_HEIGHT = 18
PDF_META_LINE_HEIGHT = 11
PDF_BOOKING_REFERENCES_PER_LINE = 6
PDF_SUMMARY_TOP_SPACING = 10
PDF_SUMMARY_FINAL_SPACING = 4
PDF_PAYMENT_TOP_SPACING = 30
PDF_PAYMENT_DEBIT_BOX_HEIGHT = 65
PDF_PAYMENT_CREDIT_BOX_HEIGHT = 70


def safe_csv_cell(value: Any) -> Any:
    """Return a CSV cell value that spreadsheet apps cannot interpret as a formula."""
    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(CSV_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def safe_csv_row(row: Iterable[Any]) -> list[Any]:
    """Return a CSV row with formula-like text cells escaped."""
    return [safe_csv_cell(value) for value in row]


def csv_response(filename, rows, headers):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(safe_csv_row(headers))
    writer.writerows(safe_csv_row(row) for row in rows)
    response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def settlement_run_csv_bytes(run: SettlementRun) -> bytes:
    """Render a versioned settlement run CSV from immutable snapshots."""
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(safe_csv_row(["Teilnehmer", "Brutto", "Förderung", "Soll", "Gezahlt", "Vorgestreckt", "Offen"]))
    for snapshot in run.settlements.all():
        writer.writerow(
            safe_csv_row(
                [
                    snapshot.participant_name,
                    snapshot.total_gross,
                    snapshot.total_subsidy,
                    snapshot.total_due,
                    snapshot.total_paid,
                    snapshot.total_advanced,
                    snapshot.balance,
                ]
            )
        )
    return buffer.getvalue().encode("utf-8")


def _format_money_cells(sheet, row: int, columns: tuple[int, ...]) -> None:
    for column in columns:
        sheet.cell(row=row, column=column).number_format = "#,##0.00"


def _decimal_text(value: str | Decimal) -> float:
    return float(Decimal(str(value)))


def _write_cost_center_sheet_from_snapshot(sheet, cost_centers: list[dict]) -> None:
    sheet.append(["Kostenstelle", "Einnahmen", "Ausgaben", "Saldo", "Anzahl Einnahmen", "Anzahl Ausgaben"])
    for data in cost_centers:
        sheet.append(
            [
                data["label"],
                _decimal_text(data["income"]),
                _decimal_text(data["expense_total"]),
                _decimal_text(data["balance"]),
                data["income_count"],
                data["expense_count"],
            ]
        )
        _format_money_cells(sheet, sheet.max_row, (2, 3, 4))

    sheet.append([])
    sheet.append(["Detaillierte Einnahmen pro Kostenstelle"])
    sheet.append(["Kostenstelle", "Datum", "Teilnehmer", "Beschreibung", "Betrag"])
    for data in cost_centers:
        for income in data["income_details"]:
            participant_name = income["participant_name"]
            if income.get("family_member_name"):
                participant_name = f"{participant_name} für {income['family_member_name']}"
            sheet.append(
                [
                    data["label"],
                    income["meal_date"][8:10] + "." + income["meal_date"][5:7] + "." + income["meal_date"][:4],
                    participant_name,
                    income["description"],
                    _decimal_text(income["amount"]),
                ]
            )
            _format_money_cells(sheet, sheet.max_row, (5,))

    sheet.append([])
    sheet.append(["Detaillierte Ausgaben pro Kostenstelle"])
    sheet.append(["Kostenstelle", "Datum", "Antragsteller", "Beschreibung", "Betrag"])
    for data in cost_centers:
        for expense in data["expense_details"]:
            sheet.append(
                [
                    data["label"],
                    expense["paid_date"][8:10] + "." + expense["paid_date"][5:7] + "." + expense["paid_date"][:4],
                    expense["applicant_name"],
                    expense["description"],
                    _decimal_text(expense["amount"]),
                ]
            )
            _format_money_cells(sheet, sheet.max_row, (5,))


def participant_import_template_response():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Teilnehmer"

    headers = [
        "Vorname*",
        "Nachname*",
        "Anreise*",
        "Abreise*",
        "Hilfssatz*",
        "Berufssatz*",
        "Email",
        "Telefon",
        "Status",
        "Kind",
        "Jugendgruppe",
        "Begleitperson",
        "Notizen",
    ]
    sheet.append(headers)

    # Example 1: Standard Active Adult
    sheet.append(
        [
            "Max",
            "Mustermann",
            "01.08.2026",
            "10.08.2026",
            1.0,
            1.0,
            "max@example.com",
            "015112345678",
            "active",
            "Nein",
            "Nein",
            "Nein",
            "Standard Flieger",
        ]
    )

    # Example 2: Child with youth group discount
    sheet.append(
        [
            "Lisa",
            "Müller",
            "01.08.2026",
            "10.08.2026",
            0.5,
            0.33,
            "",
            "",
            "registered",
            "Ja",
            "Ja",
            "Nein",
            "Vegetarisch",
        ]
    )

    # Example 3: Companion (Begleitperson, does not fly)
    sheet.append(
        [
            "Anna",
            "Schmidt",
            "05.08.2026",
            "10.08.2026",
            0.0,
            0.0,
            "anna@example.com",
            "",
            "active",
            "Nein",
            "Nein",
            "Ja",
            "Begleitperson von Lisa",
        ]
    )

    # Example 4: Student (reduced Hilfssatz)
    sheet.append(
        [
            "Tom",
            "Schulz",
            "01.08.2026",
            "08.08.2026",
            0.5,
            0.5,
            "tom@uni.de",
            "",
            "active",
            "Nein",
            "Nein",
            "Nein",
            "Studentenrabatt",
        ]
    )

    # Make headers bold
    from openpyxl.styles import Font

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Auto-adjust column widths roughly
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="teilnehmer_import_vorlage.xlsx"'
    return response


def camp_settlement_csv(camp):
    rows = []
    for result in calculate_camp_settlements(camp):
        rows.append(
            [
                result.participant.last_name,
                result.participant.first_name,
                result.total_gross,
                result.total_subsidy,
                result.total_due,
                result.total_paid,
                result.total_advanced,
                result.balance,
            ]
        )
    return csv_response(
        f"abrechnung-{camp.year}.csv",
        rows,
        ["Nachname", "Vorname", "Brutto", "Förderung", "Soll", "Gezahlt", "Vorgestreckt", "Offen"],
    )


def drink_entries_csv(camp):
    rows = []
    legacy_entries = DrinkEntry.objects.filter(participant__camp=camp).select_related("participant")
    kiosk_charges = Charge.objects.filter(
        participant__camp=camp,
        kind=Charge.Kind.DRINK,
        deleted_at__isnull=True,
    ).select_related("participant")
    for entry in legacy_entries:
        rows.append(
            [
                entry.participant.last_name,
                entry.participant.first_name,
                entry.get_drink_display(),
                entry.quantity,
                entry.unit_price,
                entry.total,
                entry.booked_at,
            ]
        )
    for entry in kiosk_charges:
        rows.append(
            [
                entry.participant.last_name,
                entry.participant.first_name,
                entry.description,
                entry.quantity,
                entry.unit_price,
                money(entry.total),
                entry.created_at,
            ]
        )
    return csv_response(
        f"getraenke-{camp.year}.csv",
        rows,
        ["Nachname", "Vorname", "Getränk", "Menge", "Einzelpreis", "Summe", "Erfasst am"],
    )


def camp_workbook_response(camp):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Abrechnung"
    summary.append(["Nachname", "Vorname", "Brutto", "Förderung", "Soll", "Gezahlt", "Vorgestreckt", "Offen"])
    for result in calculate_camp_settlements(camp):
        summary.append(
            [
                result.participant.last_name,
                result.participant.first_name,
                result.total_gross,
                result.total_subsidy,
                result.total_due,
                result.total_paid,
                result.total_advanced,
                result.balance,
            ]
        )

    participants = workbook.create_sheet("Teilnehmer")
    participants.append(["Nachname", "Vorname", "E-Mail", "Telefon", "Status", "Hilfssatz", "Berufssatz", "Nächte"])
    for participant in Participant.objects.filter(camp=camp):
        participants.append(
            [
                participant.last_name,
                participant.first_name,
                participant.email,
                participant.phone,
                participant.get_status_display(),
                participant.hilfssatz,
                participant.berufssatz,
                participant.actual_nights,
            ]
        )

    cost_centers_sheet = workbook.create_sheet("Kostenstellen")
    cost_centers = get_cost_center_evaluation(camp)
    _write_cost_center_sheet_from_snapshot(
        cost_centers_sheet,
        [
            {
                "label": data["label"],
                "income": data["income"],
                "expense_total": data["expense_total"],
                "balance": data["balance"],
                "income_count": data["income_count"],
                "expense_count": data["expense_count"],
                "income_details": [
                    {
                        "meal_date": signup.meal_date.isoformat(),
                        "participant_name": signup.participant.full_name,
                        "family_member_name": signup.family_member.full_name if signup.family_member else "",
                        "description": signup.get_meal_display(),
                        "amount": signup.charge.total,
                    }
                    for signup in data["income_details"]
                ],
                "expense_details": [
                    {
                        "paid_date": (exp.paid_on or exp.created_at.date()).isoformat(),
                        "applicant_name": exp.participant.full_name if exp.participant else "Unbekannt",
                        "description": exp.description,
                        "amount": exp.amount,
                    }
                    for exp in data["expense_details"]
                ],
            }
            for data in cost_centers.values()
        ],
    )

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="fliegerlager-{camp.year}.xlsx"'
    return response


def _draw_page_framework(pdf, title, subtitle, participant_name):
    width, height = A4

    logo_path = settings.BASE_DIR / "static" / "billing" / "logo.jpg"
    if logo_path.exists():
        pdf.drawImage(
            str(logo_path), 50, height - 150, width=250, height=100, preserveAspectRatio=True, anchor="nw", mask="auto"
        )

    pdf.setFont("Helvetica", 8)
    pdf.setFillColorRGB(0.3, 0.3, 0.3)
    pdf.drawString(50, height - 165, "Luftsportfreunde Wesel-Rheinhausen e.V. · Postfach 100240 · 46462 Wesel")
    pdf.setFillColorRGB(0, 0, 0)

    pdf.setFont("Helvetica", 10)
    pdf.drawString(50, height - 200, "An:")
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, height - 215, participant_name)

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawRightString(width - 50, height - 70, title)

    if subtitle:
        pdf.setFont("Helvetica", 10)
        pdf.setFillColorRGB(0.3, 0.3, 0.3)
        pdf.drawRightString(width - 50, height - 90, subtitle)
        pdf.setFillColorRGB(0, 0, 0)

    y = height - 260

    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(50, y - 6, width - 100, 20, stroke=0, fill=1)
    pdf.setFillColorRGB(0, 0, 0)

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(55, y, "Position")
    pdf.drawRightString(width - 120, y, "Menge")
    pdf.drawRightString(width - 55, y, "Summe")
    y -= 15

    footer_y = 30
    pdf.setFont("Helvetica", 8)
    pdf.setFillColorRGB(0.5, 0.5, 0.5)
    pdf.drawCentredString(
        width / 2.0, footer_y, "Erstellt mit der Fliegerlagerabrechnung | Luftsportfreunde Wesel-Rheinhausen e.V."
    )
    pdf.drawRightString(width - 50, footer_y, f"Seite {pdf.getPageNumber()}")
    pdf.setFillColorRGB(0, 0, 0)

    return y


def _ensure_invoice_space(pdf, y, required_height, title, subtitle, participant_name):
    """Start a continuation page when a complete invoice block would cross the footer."""
    if y - required_height >= PDF_CONTENT_BOTTOM:
        return y

    pdf.showPage()
    y = _draw_page_framework(pdf, title, subtitle, participant_name)
    pdf.setFont("Helvetica", 10)
    return y


def _invoice_line_height(occurred_on: Any, booking_references: Sequence[str]) -> int:
    metadata_lines = int(bool(occurred_on))
    if booking_references:
        metadata_lines += (len(booking_references) + PDF_BOOKING_REFERENCES_PER_LINE - 1) // (
            PDF_BOOKING_REFERENCES_PER_LINE
        )
    return PDF_LINE_HEIGHT + (metadata_lines * PDF_META_LINE_HEIGHT)


def _format_invoice_date(value: Any) -> str:
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    try:
        year, month, day = str(value).split("-", maxsplit=2)
    except ValueError:
        return str(value)
    return f"{day}.{month}.{year}"


def _draw_invoice_line(
    pdf,
    y: float,
    *,
    label: str,
    quantity: Any,
    total: Decimal,
    occurred_on: Any,
    booking_references: Sequence[str],
) -> float:
    width, _ = A4
    pdf.setFont("Helvetica", 10)
    pdf.drawString(50, y, label[:80])
    pdf.drawRightString(width - 120, y, str(quantity))
    total_str = f"- {total:.2f} €" if total > 0 else f"{total:.2f} €"
    pdf.drawRightString(width - 50, y, total_str)

    metadata_y = y - PDF_META_LINE_HEIGHT
    formatted_date = _format_invoice_date(occurred_on)
    metadata_lines = []
    if formatted_date:
        metadata_lines.append(f"Datum: {formatted_date}")
    for offset in range(0, len(booking_references), PDF_BOOKING_REFERENCES_PER_LINE):
        references = booking_references[offset : offset + PDF_BOOKING_REFERENCES_PER_LINE]
        prefix = "Buchungen: " if offset == 0 else ""
        metadata_lines.append(f"{prefix}{', '.join(references)}")

    pdf.setFillColorRGB(0.35, 0.35, 0.35)
    for metadata_line in metadata_lines:
        pdf.setFont("Helvetica", 8)
        pdf.drawString(55, metadata_y, metadata_line)
        metadata_y -= PDF_META_LINE_HEIGHT
    pdf.setFillColorRGB(0, 0, 0)

    next_y = y - _invoice_line_height(occurred_on, booking_references)
    pdf.setStrokeColorRGB(0.9, 0.9, 0.9)
    pdf.line(50, next_y + 10, width - 50, next_y + 10)
    pdf.setStrokeColorRGB(0, 0, 0)
    pdf.setFont("Helvetica", 10)
    return next_y


def _sum_block_height(items: Sequence[tuple[str, Decimal]]) -> int:
    final_spacing = PDF_SUMMARY_FINAL_SPACING if any(label == "Offen" for label, _value in items) else 0
    return PDF_SUMMARY_TOP_SPACING + (len(items) * PDF_LINE_HEIGHT) + final_spacing


def _draw_sum_block(pdf, y, items):
    width, _ = A4
    y -= PDF_SUMMARY_TOP_SPACING

    pdf.setStrokeColorRGB(0.5, 0.5, 0.5)
    pdf.line(width - 250, y + 16, width - 50, y + 16)
    pdf.setStrokeColorRGB(0, 0, 0)

    for label, value in items:
        if label in ["Brutto", "Soll"]:
            val_str = f"- {value:.2f} €" if value > 0 else f"{value:.2f} €"
        elif label in ["Förderung", "Gezahlt", "Vorgestreckt"]:
            val_str = f"+ {value:.2f} €" if value > 0 else f"{value:.2f} €"
        elif label == "Offen":
            label = "Kontostand"
            if value > 0:
                val_str = f"- {value:.2f} €"
            elif value < 0:
                val_str = f"+ {abs(value):.2f} €"
            else:
                val_str = "0.00 €"
        else:
            val_str = f"{value:.2f} €"

        is_final = label == "Kontostand"

        if is_final:
            y -= PDF_SUMMARY_FINAL_SPACING
            pdf.setStrokeColorRGB(0.2, 0.2, 0.2)
            pdf.line(width - 250, y + 14, width - 50, y + 14)
            pdf.setStrokeColorRGB(0, 0, 0)
            pdf.setFont("Helvetica-Bold", 12)
        else:
            pdf.setFont("Helvetica", 11)

        pdf.drawString(width - 220, y, f"{label}:")
        pdf.drawRightString(width - 50, y, val_str)
        y -= PDF_LINE_HEIGHT

        if is_final:
            pdf.setStrokeColorRGB(0.2, 0.2, 0.2)
            pdf.line(width - 250, y + 14, width - 50, y + 14)
            pdf.setStrokeColorRGB(0, 0, 0)

    return y


def _payment_instructions_height(camp, balance) -> int:
    if balance == 0:
        return 0
    if balance < 0:
        return PDF_PAYMENT_TOP_SPACING + PDF_PAYMENT_CREDIT_BOX_HEIGHT

    iban = getattr(camp, "iban", "").strip()
    paypal = getattr(camp, "paypal_link", "").strip()
    if not iban and not paypal:
        return 0
    return PDF_PAYMENT_TOP_SPACING + PDF_PAYMENT_DEBIT_BOX_HEIGHT


def _draw_payment_instructions(pdf, y, camp, balance):
    if balance == 0:
        return y

    width, _ = A4

    if balance > 0:
        iban = getattr(camp, "iban", "").strip()
        paypal = getattr(camp, "paypal_link", "").strip()

        if not iban and not paypal:
            return y

        box_height = PDF_PAYMENT_DEBIT_BOX_HEIGHT
        y -= PDF_PAYMENT_TOP_SPACING
        y -= box_height

        pdf.setFillColorRGB(0.96, 0.96, 0.96)
        pdf.setStrokeColorRGB(0.85, 0.85, 0.85)
        pdf.roundRect(50, y, width - 100, box_height, radius=4, stroke=1, fill=1)

        text_y = y + box_height - 18
        pdf.setFillColorRGB(0, 0, 0)
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(65, text_y, "Zahlungsinformationen")

        text_y -= 14
        pdf.setFont("Helvetica", 9)
        pdf.drawString(65, text_y, "Bitte begleiche den offenen Kontostand zeitnah auf eines der folgenden Konten:")

        text_y -= 16
        pdf.setFont("Helvetica-Bold", 9)

        if iban and paypal:
            pdf.drawString(65, text_y, f"IBAN: {iban}")
            pdf.drawString(280, text_y, f"PayPal: {paypal}")
        elif iban:
            pdf.drawString(65, text_y, f"IBAN: {iban}")
        elif paypal:
            pdf.drawString(65, text_y, f"PayPal: {paypal}")

    else:
        # balance < 0 (Guthaben)
        box_height = PDF_PAYMENT_CREDIT_BOX_HEIGHT
        y -= PDF_PAYMENT_TOP_SPACING
        y -= box_height

        pdf.setFillColorRGB(0.96, 0.96, 0.96)
        pdf.setStrokeColorRGB(0.85, 0.85, 0.85)
        pdf.roundRect(50, y, width - 100, box_height, radius=4, stroke=1, fill=1)

        text_y = y + box_height - 18
        pdf.setFillColorRGB(0, 0, 0)
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(65, text_y, "Guthaben & Auszahlung")

        text_y -= 14
        pdf.setFont("Helvetica", 9)
        pdf.drawString(
            65, text_y, "Du hast ein Guthaben. Bitte teile der Lagerleitung mit, ob du diesen Betrag spenden (auch"
        )
        text_y -= 12
        pdf.drawString(65, text_y, "anteilig möglich) oder ausgezahlt haben möchtest.")

        text_y -= 16
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(65, text_y, "Für eine Auszahlung nenne der Lagerleitung bitte deine IBAN oder PayPal-Adresse.")

    return y


def participant_pdf_response(participant):
    result = calculate_participant_settlement(participant)
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, _ = A4

    title = f"Einzelabrechnung {participant.camp.name} {participant.camp.year}"
    y = _draw_page_framework(pdf, title, "", participant.full_name)

    pdf.setFont("Helvetica", 10)
    for line in result.lines:
        line_height = _invoice_line_height(line.occurred_on, line.booking_references)
        y = _ensure_invoice_space(
            pdf,
            y,
            line_height,
            title,
            "",
            participant.full_name,
        )
        y = _draw_invoice_line(
            pdf,
            y,
            label=line.label,
            quantity=line.quantity,
            total=line.total,
            occurred_on=line.occurred_on,
            booking_references=line.booking_references,
        )

    summary_items = [
        ("Brutto", result.total_gross),
        ("Förderung", result.total_subsidy),
        ("Soll", result.total_due),
        ("Gezahlt", result.total_paid),
        ("Vorgestreckt", result.total_advanced),
        ("Offen", result.balance),
    ]
    closing_height = _sum_block_height(summary_items) + _payment_instructions_height(
        participant.camp,
        result.balance,
    )
    y = _ensure_invoice_space(pdf, y, closing_height, title, "", participant.full_name)
    y = _draw_sum_block(pdf, y, summary_items)

    _draw_payment_instructions(pdf, y, participant.camp, result.balance)

    pdf.showPage()
    pdf.save()

    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="abrechnung-{participant.pk}.pdf"'
    return response


def settlement_run_csv(run: SettlementRun) -> HttpResponse:
    response = HttpResponse(settlement_run_csv_bytes(run), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="abrechnung-{run.camp.year}-v{run.version}.csv"'
    return response


def settlement_run_workbook_bytes(run: SettlementRun) -> bytes:
    """Render a versioned settlement run workbook from immutable snapshots."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Abrechnung"
    summary.append(["Teilnehmer", "Status", "Brutto", "Förderung", "Soll", "Gezahlt", "Vorgestreckt", "Offen"])
    for snapshot in run.settlements.all():
        summary.append(
            [
                snapshot.participant_name,
                snapshot.data.get("participant", {}).get("status_label", snapshot.participant_status),
                snapshot.total_gross,
                snapshot.total_subsidy,
                snapshot.total_due,
                snapshot.total_paid,
                snapshot.total_advanced,
                snapshot.balance,
            ]
        )

    cost_centers_sheet = workbook.create_sheet("Kostenstellen")
    _write_cost_center_sheet_from_snapshot(cost_centers_sheet, run.cost_center_data or [])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def settlement_run_workbook_response(run: SettlementRun) -> HttpResponse:
    response = HttpResponse(
        settlement_run_workbook_bytes(run),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="abrechnung-{run.camp.year}-v{run.version}.xlsx"'
    return response


def settlement_snapshot_pdf_bytes(snapshot: Settlement) -> bytes:
    """Render a PDF invoice from a settlement snapshot without persisting it."""
    run = snapshot.run
    if run is None:
        raise ValueError("Historical settlement PDF requires a versioned run.")
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, _ = A4

    title = f"Einzelabrechnung {run.camp.name} {run.camp.year}"
    subtitle = f"Version {run.version} vom {run.created_at:%d.%m.%Y %H:%M}"
    y = _draw_page_framework(pdf, title, subtitle, snapshot.participant_name)

    pdf.setFont("Helvetica", 10)
    for line in snapshot.data.get("lines", []):
        booking_references = tuple(line.get("booking_references") or ())
        occurred_on = line.get("occurred_on")
        line_height = _invoice_line_height(occurred_on, booking_references)
        y = _ensure_invoice_space(
            pdf,
            y,
            line_height,
            title,
            subtitle,
            snapshot.participant_name,
        )
        try:
            total = Decimal(str(line.get("total", "0.00")))
        except (ValueError, TypeError):
            total = Decimal("0.00")
        y = _draw_invoice_line(
            pdf,
            y,
            label=str(line.get("label", "")),
            quantity=str(line.get("quantity", "")),
            total=total,
            occurred_on=occurred_on,
            booking_references=booking_references,
        )

    summary_items = [
        ("Brutto", snapshot.total_gross),
        ("Förderung", snapshot.total_subsidy),
        ("Soll", snapshot.total_due),
        ("Gezahlt", snapshot.total_paid),
        ("Vorgestreckt", snapshot.total_advanced),
        ("Offen", snapshot.balance),
    ]
    closing_height = _sum_block_height(summary_items) + _payment_instructions_height(
        run.camp,
        snapshot.balance,
    )
    y = _ensure_invoice_space(pdf, y, closing_height, title, subtitle, snapshot.participant_name)
    y = _draw_sum_block(pdf, y, summary_items)

    _draw_payment_instructions(pdf, y, run.camp, snapshot.balance)

    pdf.showPage()
    pdf.save()
    return output.getvalue()


def settlement_snapshot_pdf_response(snapshot: Settlement) -> HttpResponse:
    run = snapshot.run
    if run is None:
        raise ValueError("Historical settlement PDF requires a versioned run.")
    response = HttpResponse(settlement_snapshot_pdf_bytes(snapshot), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="abrechnung-{snapshot.pk}-v{run.version}.pdf"'
    return response
