import csv
from dataclasses import asdict, dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO, TextIOWrapper
from zipfile import BadZipFile

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import Participant

XLSX_MAGIC = b"PK\x03\x04"
MAX_IMPORT_ROWS = 5000
REQUIRED_PARTICIPANT_COLUMNS = ["first_name", "last_name"]
PARTICIPANT_COLUMNS = [
    "first_name",
    "last_name",
    "email",
    "phone",
    "status",
    "is_child",
    "is_youth_group",
    "is_companion",
    "hilfssatz",
    "berufssatz",
    "arrival_date",
    "departure_date",
    "booked_nights",
    "actual_nights",
    "notes",
]


def parse_date(value, field_name, errors):
    if not value or not str(value).strip():
        return None
    val_str = str(value).strip()
    import datetime
    
    # Try different date formats
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    errors.append(f"{field_name}: ungültiges Datumsformat (erwartet TT.MM.JJJJ)")
    return None


@dataclass
class ImportRow:
    row_number: int
    data: dict
    errors: list[str] = field(default_factory=list)

    @property
    def valid(self):
        return not self.errors


def parse_bool(value):
    if value in (True, False):
        return bool(value)
    if value is None or value == "":
        return False
    return str(value).strip().lower() in {"1", "true", "ja", "yes", "x"}


def parse_int(value, field_name, errors):
    if value in (None, ""):
        return 0
    try:
        return int(Decimal(str(value).replace(",", ".")))
    except (InvalidOperation, ValueError):
        errors.append(f"{field_name}: keine gültige Zahl")
        return 0


def parse_decimal(value, field_name, errors, default="1"):
    if value in (None, ""):
        return Decimal(default)
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        errors.append(f"{field_name}: keine gültige Zahl")
        return Decimal(default)


def normalize_row(raw, row_number):
    errors = []
    
    mapping = {
        "vorname": "first_name", "first_name": "first_name",
        "nachname": "last_name", "last_name": "last_name",
        "anreise": "arrival_date", "arrival_date": "arrival_date",
        "abreise": "departure_date", "departure_date": "departure_date",
        "hilfssatz": "hilfssatz",
        "berufssatz": "berufssatz",
        "email": "email", "e-mail": "email",
        "telefon": "phone", "phone": "phone",
        "status": "status",
        "kind": "is_child", "is_child": "is_child",
        "jugendgruppe": "is_youth_group", "is_youth_group": "is_youth_group",
        "begleitperson": "is_companion", "is_companion": "is_companion",
        "notizen": "notes", "notes": "notes",
        "gebuchte_nächte": "booked_nights", "booked_nights": "booked_nights",
        "ist_nächte": "actual_nights", "actual_nights": "actual_nights",
    }
    
    data = {}
    extra_notes = []
    
    for k, v in raw.items():
        if k is None:
            continue
        k_lower = str(k).lower().replace("*", "").strip()
        # Convert None to empty string for string/text fields
        if v is None:
            v = ""
        if k_lower in mapping:
            data[mapping[k_lower]] = v
        else:
            if str(v).strip():
                extra_notes.append(f"{k}: {v}")
                
    for column in PARTICIPANT_COLUMNS:
        data.setdefault(column, "")
        
    for column in ["first_name", "last_name", "arrival_date", "departure_date", "hilfssatz", "berufssatz"]:
        if not str(data.get(column, "")).strip():
            label_map = {
                "first_name": "Vorname", "last_name": "Nachname",
                "arrival_date": "Anreise", "departure_date": "Abreise",
                "hilfssatz": "Hilfssatz", "berufssatz": "Berufssatz"
            }
            errors.append(f"{label_map.get(column, column)}: Pflichtfeld fehlt")
            
    data["arrival_date"] = parse_date(data.get("arrival_date"), "Anreise", errors)
    data["departure_date"] = parse_date(data.get("departure_date"), "Abreise", errors)

    for column in ["booked_nights", "actual_nights"]:
        data[column] = parse_int(data.get(column), column, errors)
        
    for column in ["hilfssatz", "berufssatz"]:
        val = parse_decimal(data.get(column), column, errors)
        if val is not None and (val < 0 or val > 1):
            errors.append(f"{column.capitalize()}: Wert muss zwischen 0 und 1 liegen")
        data[column] = val
        
    for column in ["is_child", "is_youth_group", "is_companion"]:
        data[column] = parse_bool(data.get(column))
        
    if extra_notes:
        if data["notes"]:
            data["notes"] += "\n" + "\n".join(extra_notes)
        else:
            data["notes"] = "\n".join(extra_notes)

    return ImportRow(row_number=row_number, data=data, errors=errors)


def read_csv(file_obj):
    try:
        wrapper = TextIOWrapper(file_obj, encoding="utf-8-sig")
        reader = csv.DictReader(wrapper)
        rows = []
        for index, row in enumerate(reader, start=2):
            if len(rows) >= MAX_IMPORT_ROWS:
                raise ValidationError(f"Die Importdatei darf höchstens {MAX_IMPORT_ROWS} Datenzeilen enthalten.")
            rows.append(normalize_row(row, index))
        return rows
    except UnicodeDecodeError as error:
        raise ValidationError("Die CSV-Datei muss UTF-8-kodiert sein.") from error


def read_xlsx(file_obj):
    try:
        workbook = load_workbook(BytesIO(file_obj.read()), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError) as error:
        raise ValidationError("Die XLSX-Datei konnte nicht sicher gelesen werden.") from error
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []
        headers = [str(value).strip() if value is not None else "" for value in header_row]
        parsed = []
        for index, values in enumerate(rows, start=2):
            if len(parsed) >= MAX_IMPORT_ROWS:
                raise ValidationError(f"Die Importdatei darf höchstens {MAX_IMPORT_ROWS} Datenzeilen enthalten.")
            raw = dict(zip(headers, values, strict=False))
            parsed.append(normalize_row(raw, index))
        return parsed
    finally:
        workbook.close()


def _peek(file_obj, size=4):
    position = file_obj.tell() if file_obj.seekable() else None
    header = file_obj.read(size)
    if position is not None:
        file_obj.seek(position)
    return header


def preview_participants(file_obj, filename):
    normalized_name = filename.lower()
    header = _peek(file_obj)
    if normalized_name.endswith(".xlsx"):
        if not header.startswith(XLSX_MAGIC):
            raise ValidationError("Die XLSX-Datei hat kein gültiges Excel-Dateiformat.")
        return read_xlsx(file_obj)
    if header.startswith(XLSX_MAGIC):
        raise ValidationError("Die Datei ist eine Excel-Datei, hat aber nicht die Endung .xlsx.")
    return read_csv(file_obj)


def rows_to_payload(rows):
    return [asdict(row) for row in rows]


def rows_from_payload(payload):
    import datetime
    
    rows = []
    for row in payload:
        data = row["data"]
        # Deserialize date strings back to date objects
        for field in ["arrival_date", "departure_date"]:
            if data.get(field) and isinstance(data[field], str):
                try:
                    data[field] = datetime.date.fromisoformat(data[field])
                except ValueError:
                    pass
                    
        # Decimal fields might also be strings or floats, though JSON handles floats. 
        # For precision, let's leave them if Django ORM handles them, or parse them to Decimal.
        for field in ["hilfssatz", "berufssatz", "quantity", "unit_price", "amount"]:
            if field in data and data[field] is not None:
                data[field] = Decimal(str(data[field]))
                
        rows.append(ImportRow(row_number=row["row_number"], data=data, errors=row.get("errors", [])))
    return rows


def save_participants(camp, rows):
    created = []
    with transaction.atomic():
        for row in rows:
            if not row.valid:
                continue
            archived = Participant.objects.filter(
                camp=camp,
                first_name=row.data["first_name"],
                last_name=row.data["last_name"],
                archived_at__isnull=False,
            ).exists()
            if archived:
                raise ValidationError(
                    f"{row.data['first_name']} {row.data['last_name']} ist archiviert "
                    "und muss zuerst wiederhergestellt werden."
                )
            participant, _ = Participant.objects.update_or_create(
                camp=camp,
                first_name=row.data["first_name"],
                last_name=row.data["last_name"],
                defaults={
                    field: row.data[field] for field in PARTICIPANT_COLUMNS if field not in {"first_name", "last_name"}
                },
            )
            created.append(participant)
    return created
