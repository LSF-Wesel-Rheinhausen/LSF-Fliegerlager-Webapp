"""Create the administrator attendance workbook for one camp."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from django.core.exceptions import ValidationError
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .attendance import attendance_window, iter_overnight_dates, target_stay_for
from .models import AttendanceDay, Participant, ParticipantFamilyMember

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_ACTIVE_PARTICIPANT_STATUSES = (
    Participant.Status.REGISTERED,
    Participant.Status.ACTIVE,
    Participant.Status.SETTLED,
)
_STATUS_PRESENT = "AN"
_STATUS_ABSENT = "AB"
_STATUS_NOT_RELEVANT = "–"
_PERSON_TYPE_PARTICIPANT = "Teilnehmer"
_PERSON_TYPE_CHILD = "Kind"
_PERSON_TYPE_COMPANION = "Begleitperson"
_STATUS_STYLES = {
    _STATUS_PRESENT: ("Anwesend", "FFC6EFCE", "FF006100"),
    _STATUS_ABSENT: ("Abwesend", "FFFFC7CE", "FF9C0006"),
    _STATUS_NOT_RELEVANT: ("Außerhalb des Aufenthaltszeitraums", "FFE7E6E6", "FF595959"),
}


def _safe_cell_value(value: Any) -> Any:
    """Return a value that cannot be interpreted as a spreadsheet formula."""
    if isinstance(value, str) and value.lstrip().startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _camp_dates(camp: Any) -> list[date]:
    window = attendance_window(camp)
    if window is None:
        return []
    return list(iter_overnight_dates(*window))


def _age_on(person: Any, reference_date: date) -> int | None:
    try:
        return person.age_on(reference_date)
    except ValidationError:
        return None


def _in_range(person: Any, camp: Any, day: date) -> bool:
    stay = target_stay_for(person, camp)
    return stay is not None and stay[0] <= day < stay[1]


def _person_type(person: Participant | ParticipantFamilyMember) -> str:
    if isinstance(person, ParticipantFamilyMember):
        if person.role == ParticipantFamilyMember.Role.CHILD:
            return _PERSON_TYPE_CHILD
        return _PERSON_TYPE_COMPANION
    if person.is_child:
        return _PERSON_TYPE_CHILD
    if person.is_companion:
        return _PERSON_TYPE_COMPANION
    return _PERSON_TYPE_PARTICIPANT


def _status(
    person: Any,
    camp: Any,
    day: date,
    attendance_by_person: dict[tuple[str, int], dict[date, AttendanceDay]],
) -> str:
    if not _in_range(person, camp, day):
        return _STATUS_NOT_RELEVANT
    if not person.attendance_tracking_enabled:
        return _STATUS_PRESENT
    key = ("family" if isinstance(person, ParticipantFamilyMember) else "participant", person.pk)
    attendance = attendance_by_person.get(key, {}).get(day)
    return _STATUS_PRESENT if attendance is not None and attendance.is_present else _STATUS_ABSENT


def _style_status_cell(cell: Any) -> None:
    status_style = _STATUS_STYLES.get(cell.value)
    if status_style is None:
        return
    _, fill_color, font_color = status_style
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(color=font_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_sheet(sheet: Any, widths: dict[int, float]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "D2" if sheet.title == "Anwesenheit" else "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _configure_attendance_printing(sheet: Any) -> None:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.scale = None
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_area = f"$A$1:${get_column_letter(sheet.max_column)}${sheet.max_row}"
    sheet.print_title_rows = "$1:$1"
    sheet.print_title_cols = "$A:$C"


def attendance_workbook_response(camp: Any) -> HttpResponse:
    """Build an XLSX download containing attendance matrix and daily totals.

    Parameters:
        camp: Camp-like object with ``year``, ``starts_on`` and ``ends_on``.

    Returns:
        An attachment response named ``anwesenheit-<year>.xlsx``.
    """
    days = _camp_dates(camp)
    workbook = Workbook()
    workbook.iso_dates = True
    attendance_sheet = workbook.active
    attendance_sheet.title = "Anwesenheit"
    summary_sheet = workbook.create_sheet("Tagesübersicht")
    legend_sheet = workbook.create_sheet("Legende")

    attendance_sheet.append(["Name", "Alter", "Typ", *days])
    summary_sheet.append(["Datum", "Anwesend", "Abwesend", "Kommentare"])
    legend_sheet.append(["Status", "Bedeutung"])
    for status, (meaning, _, _) in _STATUS_STYLES.items():
        legend_sheet.append([status, meaning])

    if days:
        participants = list(
            Participant.objects.filter(
                camp=camp,
                archived_at__isnull=True,
                status__in=_ACTIVE_PARTICIPANT_STATUSES,
            ).prefetch_related("family_members")
        )
        people: list[Participant | ParticipantFamilyMember] = list(participants)
        for participant in participants:
            people.extend(member for member in participant.family_members.all() if member.is_active)
        people.sort(key=lambda person: (person.last_name.casefold(), person.first_name.casefold(), person.pk))

        attendance_records = list(
            AttendanceDay.objects.filter(participant__camp=camp).select_related("family_member").order_by("date", "pk")
        )
        attendance_by_person: dict[tuple[str, int], dict[date, AttendanceDay]] = {}
        for attendance in attendance_records:
            key = (
                "family" if attendance.family_member_id is not None else "participant",
                attendance.family_member_id or attendance.participant_id,
            )
            attendance_by_person.setdefault(key, {})[attendance.date] = attendance

        for person in people:
            attendance_sheet.append(
                [
                    _safe_cell_value(person.full_name),
                    _age_on(person, camp.starts_on),
                    _safe_cell_value(_person_type(person)),
                    *[_status(person, camp, day, attendance_by_person) for day in days],
                ]
            )

        for day in days:
            statuses = [_status(person, camp, day, attendance_by_person) for person in people]
            comments: list[str] = []
            people_by_key = {
                ("family" if isinstance(person, ParticipantFamilyMember) else "participant", person.pk): person
                for person in people
            }
            for attendance in attendance_records:
                if attendance.date != day or not attendance.comment:
                    continue
                key = (
                    "family" if attendance.family_member_id is not None else "participant",
                    attendance.family_member_id or attendance.participant_id,
                )
                target_person = people_by_key.get(key)
                if target_person is not None and _in_range(target_person, camp, day):
                    comments.append(_safe_cell_value(attendance.comment))
            summary_sheet.append(
                [
                    day,
                    statuses.count(_STATUS_PRESENT),
                    statuses.count(_STATUS_ABSENT),
                    "; ".join(comments),
                ]
            )

    for column in range(4, attendance_sheet.max_column + 1):
        attendance_sheet.cell(1, column).number_format = "DD.MM.YYYY"
    for row in range(2, summary_sheet.max_row + 1):
        summary_sheet.cell(row, 1).number_format = "DD.MM.YYYY"
    for row in attendance_sheet.iter_rows(
        min_row=2,
        min_col=4,
        max_row=attendance_sheet.max_row,
        max_col=attendance_sheet.max_column,
    ):
        for cell in row:
            _style_status_cell(cell)
    for row in range(2, legend_sheet.max_row + 1):
        _style_status_cell(legend_sheet.cell(row, 1))

    _style_sheet(
        attendance_sheet,
        {1: 30, 2: 10, 3: 16, **{column: 6 for column in range(4, attendance_sheet.max_column + 1)}},
    )
    _style_sheet(summary_sheet, {1: 14, 2: 12, 3: 12, 4: 60})
    _style_sheet(legend_sheet, {1: 10, 2: 32})
    _configure_attendance_printing(attendance_sheet)

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="anwesenheit-{camp.year}.xlsx"'
    return response
